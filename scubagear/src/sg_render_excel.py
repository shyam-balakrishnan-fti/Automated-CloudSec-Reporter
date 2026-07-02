"""
sg_render_excel.py — Stage 5: Excel output renderer

Populates the Azure sheet in Output_Template.xlsx with enriched M365 findings.

Template structure (Azure sheet):
  Row 1:  Column headers (Ref, Finding, Risk Rating, ...)
  Row 2+: Section heading rows (bold, col A only) interleaved with finding rows

Section order (matches template):
  - Microsoft Entra ID (previously Azure Active Directory)
  - Microsoft 365 Defender
  - Microsoft Exchange Online
  - Microsoft Power Platform
  - Azure Resources   ← kept in template but left empty for ScubaGear output (Option A)

Renderer writes INTO the template file by clearing the sample rows (3-25)
and re-inserting findings under the correct section headings.
"""

from __future__ import annotations

import copy
import shutil
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Optional

from sg_models import SECTION_ORDER
from sg_grouping import GroupedOutputGroup, GroupingResult
from sg_enrich import EnrichResult

try:
    import openpyxl
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    raise ImportError("openpyxl not installed. Run: pip install openpyxl")

# ── Column layout (1-indexed, matches template header row) ────────────

COL_REF         = 1
COL_FINDING     = 2
COL_RISK        = 3
COL_ROOT_CAUSE  = 4
COL_LIKELIHOOD  = 5
COL_CONSEQUENCE = 6
COL_ACCESS      = 7
COL_SITUATION   = 8
COL_CONSEQUENCE_NARR = 9
COL_RECOMMENDATIONS  = 10

# ── Style constants ───────────────────────────────────────────────────

FONT_FAMILY = "Calibri"

# Risk rating → fill colour (ARGB, matches template sample rows)
RISK_FILLS = {
    "High":   "FFFF0000",  # red
    "Medium": "FFFFC000",  # amber
    "Low":    "FF92D050",  # green
}

# Consequence rating colours (subtle)
CONSEQUENCE_FILLS = {
    "Major":    "FFFFD7D7",
    "Moderate": "FFFFF2CC",
    "Minor":    "FFE2EFDA",
}


def _risk_fill(risk: str) -> Optional[PatternFill]:
    argb = RISK_FILLS.get(risk)
    if argb:
        return PatternFill("solid", fgColor=argb)
    return None


def _wrap_align(horizontal: str = "left") -> Alignment:
    return Alignment(horizontal=horizontal, vertical="top", wrap_text=True)


def _bold_font(size: int = 11) -> Font:
    return Font(name=FONT_FAMILY, bold=True, size=size)


def _normal_font(size: int = 11) -> Font:
    return Font(name=FONT_FAMILY, bold=False, size=size)


# ── Section heading writer ────────────────────────────────────────────

def _write_section_heading(ws: Any, row: int, section_name: str) -> None:
    """Write a bold section heading in col A, matching the template style."""
    cell = ws.cell(row=row, column=COL_REF)
    cell.value     = section_name
    cell.font      = _bold_font(11)
    cell.alignment = _wrap_align()
    ws.row_dimensions[row].height = 26.4


# ── Finding row writer ────────────────────────────────────────────────

def _write_finding_row(ws: Any, row: int, group: GroupedOutputGroup) -> None:
    """Write one enriched finding group as a single row in the report."""
    rep = group.representative

    ref_label       = rep.ref_label()
    finding_title   = group.finding_title or rep.finding_title or rep.check_title or group.group_name
    risk_rating     = group.risk_rating or rep.risk_rating or "Medium"
    root_cause      = group.root_cause_narrative or rep.root_cause_narrative or ""
    likelihood      = group.likelihood_rating or rep.likelihood_rating or "Medium"
    consequence_rat = group.consequence_rating or rep.consequence_rating or "Moderate"
    access_req      = group.access_required or rep.access_required or ""
    situation       = group.situation_narrative or rep.situation_narrative or ""
    consequence_nar = group.consequence_narrative or rep.consequence_narrative or ""

    recommendations = group.recommendations or rep.recommendations or root_cause

    cells = {
        COL_REF:             ref_label,
        COL_FINDING:         finding_title,
        COL_RISK:            risk_rating,
        COL_ROOT_CAUSE:      root_cause,
        COL_LIKELIHOOD:      likelihood,
        COL_CONSEQUENCE:     consequence_rat,
        COL_ACCESS:          access_req,
        COL_SITUATION:       situation,
        COL_CONSEQUENCE_NARR: consequence_nar,
        COL_RECOMMENDATIONS: recommendations,
    }

    for col_idx, value in cells.items():
        cell           = ws.cell(row=row, column=col_idx)
        cell.value     = value or ""
        cell.font      = _normal_font(11)
        cell.alignment = _wrap_align()
        # Explicitly clear fill on non-risk columns so template ghost fills
        # from the original section-heading rows don't bleed through.
        if col_idx != COL_RISK:
            cell.fill = PatternFill(fill_type=None)

    # Risk rating cell gets colour fill
    risk_fill = _risk_fill(risk_rating)
    if risk_fill:
        ws.cell(row=row, column=COL_RISK).fill = risk_fill

    ws.row_dimensions[row].height = 60  # default row height; Excel auto-adjusts on open


# ── Ref number reassignment ──────────────────────────────────────────

def _reassign_ref_numbers(groups: list[GroupedOutputGroup]) -> None:
    """
    Reassign per-section sequential ref numbers to the final approved groups.

    Must run after grouping/review — the Stage 2 assignment is on individual
    OutputGroups and becomes stale once the analyst merges or reorders groups.
    Groups must already be sorted in section order before this is called.

    Counter resets per section: ENT1, ENT2, ..., DEF1, DEF2, ...
    """
    from collections import defaultdict
    counters: dict[str, int] = defaultdict(int)
    for g in groups:
        if not g.representative:
            continue
        section = g.output_section
        counters[section] += 1
        g.representative.ref_number = counters[section]
        # ref_prefix comes from the representative's own field (set at ingest)
        # but for merged groups the representative may come from any control —
        # use the section's expected prefix from the first source group instead.
        if g.source_groups:
            g.representative.ref_prefix = g.source_groups[0].representative.ref_prefix


def _sort_groups_for_render(groups: list[GroupedOutputGroup]) -> list[GroupedOutputGroup]:
    """
    Sort groups for Excel output:
      1. Section order (matches SECTION_ORDER / template row order)
      2. Severity within section (high → medium → low)
      3. Group name as tiebreaker for determinism
    """
    _sev = {"high": 0, "medium": 1, "low": 2}
    _sec = {s: i for i, s in enumerate(SECTION_ORDER)}
    return sorted(
        groups,
        key=lambda g: (
            _sec.get(g.output_section, 99),
            _sev.get((g.severity or "low").lower(), 3),
            g.group_name,
        ),
    )


# ── Main renderer ─────────────────────────────────────────────────────

@dataclass
class RenderResult:
    output_path: Path
    groups_written: int = 0
    sections_written: int = 0
    warnings: list[str] = field(default_factory=list)


def render_excel(
    enrich_result: EnrichResult,
    template_path: Path,
    output_path: Path,
    config: dict[str, Any],
) -> RenderResult:
    """
    Populate the Azure sheet in Output_Template.xlsx with enriched M365 findings.

    Steps:
      1. Copy template to output path (never modify the original)
      2. Open copy, target the Azure sheet
      3. Clear sample finding rows (rows 3 through end of data)
      4. Group enriched findings by output_section
      5. For each section in SECTION_ORDER: write heading row + finding rows
      6. Save

    Returns RenderResult with output path and counts.
    """
    if not template_path.exists():
        raise FileNotFoundError(f"Template not found: {template_path}")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(template_path, output_path)

    wb = load_workbook(output_path)
    target_sheet = config.get("output", {}).get("target_sheet", "Azure")

    if target_sheet not in wb.sheetnames:
        raise ValueError(
            f"Sheet '{target_sheet}' not found in template. "
            f"Available sheets: {wb.sheetnames}"
        )

    # ── Rebuild the target sheet from scratch ──────────────────────────
    # openpyxl retains named styles and cell-level formatting from the
    # template even after delete_rows — ghost cells appear wherever the
    # template had section-heading rows. The only reliable fix is to
    # snapshot the header row's column widths and row height, delete the
    # old sheet entirely, create a fresh one, then re-write the header.
    old_ws = wb[target_sheet]

    # Snapshot header row styles and column dimensions
    header_cells: list[tuple] = []  # (col_idx, value, font, fill, alignment)
    for cell in old_ws[1]:
        header_cells.append((
            cell.column,
            cell.value,
            copy.copy(cell.font) if cell.font else _bold_font(),
            copy.copy(cell.fill) if cell.fill else PatternFill(),
            copy.copy(cell.alignment) if cell.alignment else _wrap_align(),
        ))
    col_widths = {
        col: old_ws.column_dimensions[col].width
        for col in old_ws.column_dimensions
    }
    header_height = old_ws.row_dimensions[1].height or 43.2
    sheet_position = wb.sheetnames.index(target_sheet)

    # Delete and recreate
    del wb[target_sheet]
    ws = wb.create_sheet(target_sheet, sheet_position)

    # Restore column widths
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Restore header row
    ws.row_dimensions[1].height = header_height
    for col_idx, value, font, fill, alignment in header_cells:
        cell = ws.cell(row=1, column=col_idx)
        cell.value     = value
        cell.font      = font
        cell.fill      = fill
        cell.alignment = alignment

    # ── Sort groups and reassign ref numbers ─────────────────────────
    # Must happen here, after grouping/review, so merged groups and any
    # analyst reordering are reflected. Stage 2 ref numbers are stale by
    # this point because controls may have been merged across sections.
    sorted_groups = _sort_groups_for_render(list(enrich_result.output_groups))
    _reassign_ref_numbers(sorted_groups)

    # ── Group findings by output_section ──────────────────────────────
    groups_by_section: dict[str, list[GroupedOutputGroup]] = {s: [] for s in SECTION_ORDER}

    for group in sorted_groups:
        section = group.output_section
        if section not in groups_by_section:
            groups_by_section[section] = []
        groups_by_section[section].append(group)

    # ── Write sections in template order ──────────────────────────────
    current_row   = 2
    groups_written   = 0
    sections_written = 0
    warnings: list[str] = []

    for section in SECTION_ORDER:
        groups_in_section = groups_by_section.get(section, [])

        # Always write section heading (even if empty — matches template style)
        _write_section_heading(ws, current_row, section)
        current_row += 1
        sections_written += 1

        if not groups_in_section:
            # Empty section — leave heading only (Option A: leave Azure Resources etc. blank)
            continue

        for group in groups_in_section:
            if not group.representative:
                warnings.append(f"Group '{group.group_name}' has no representative — skipped")
                continue
            _write_finding_row(ws, current_row, group)
            current_row += 1
            groups_written += 1

    # ── Any findings that mapped to unknown sections ───────────────────
    unknown_groups = [
        g for g in sorted_groups
        if g.output_section not in SECTION_ORDER
    ]
    if unknown_groups:
        _write_section_heading(ws, current_row, "Other")
        current_row += 1
        for group in unknown_groups:
            if group.representative:
                _write_finding_row(ws, current_row, group)
                current_row += 1
                groups_written += 1
        warnings.append(
            f"{len(unknown_groups)} finding(s) written to 'Other' section "
            f"(unrecognised output_section value)"
        )

    wb.save(output_path)

    print(
        f"\n[ Stage 5 ] Excel rendered → {output_path.name}\n"
        f"  {groups_written} finding(s) across {sections_written} section(s)",
        flush=True,
    )
    if warnings:
        for w in warnings:
            print(f"  ⚠ {w}", flush=True)

    return RenderResult(
        output_path=output_path,
        groups_written=groups_written,
        sections_written=sections_written,
        warnings=warnings,
    )


def render_excel_from_grouping(
    grouping_result: GroupingResult,
    enrich_result: EnrichResult,
    template_path: Path,
    output_path: Path,
    config: dict[str, Any],
) -> RenderResult:
    """Convenience wrapper — same as render_excel but accepts both results."""
    return render_excel(enrich_result, template_path, output_path, config)