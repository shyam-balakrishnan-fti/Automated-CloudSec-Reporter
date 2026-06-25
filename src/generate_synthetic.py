"""
generate_synthetic.py — Synthetic Prowler data generator for Stage 1 testing.

Generates a realistic synthetic Prowler XLSX that exercises every parser code path:
    - Normal FAIL findings across S3, IAM, EC2, RDS, VPC
    - MUTED(FAIL) findings
    - PASS findings (should be excluded downstream)
    - MUTED=True with STATUS=FAIL (MUTED reconciliation case)
    - IAM/global findings with blank REGION
    - Blank DESCRIPTION/RISK (Category 2 data quality blanks)
    - Blank SUBSERVICE_NAME/RESOURCE_DETAILS (Category 1 structural blanks)
    - Multi-account findings (same check_id, different accounts)
    - Duplicate resources (same resource, same check — dedup test)
    - ARN-less findings (only RESOURCE_NAME — ARN fallback test)
    - Formula injection strings in text fields
    - Multi-value CATEGORIES and COMPLIANCE fields
    - Long DESCRIPTION field (near Excel cell limit)
    - Unknown future column (extra_fields test)
"""

from __future__ import annotations

import uuid
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

# ── Constants ─────────────────────────────────────────────────────────

HEADERS = [
    "AUTH_METHOD", "TIMESTAMP", "ACCOUNT_UID", "ACCOUNT_NAME", "ACCOUNT_EMAIL",
    "ACCOUNT_ORGANIZATION_UID", "ACCOUNT_ORGANIZATION_NAME", "ACCOUNT_TAGS",
    "FINDING_UID", "PROVIDER", "CHECK_ID", "CHECK_TITLE", "CHECK_TYPE",
    "STATUS", "STATUS_EXTENDED", "MUTED", "SERVICE_NAME", "SUBSERVICE_NAME",
    "SEVERITY", "RESOURCE_TYPE", "RESOURCE_UID", "RESOURCE_NAME",
    "RESOURCE_DETAILS", "RESOURCE_TAGS", "PARTITION", "REGION",
    "DESCRIPTION", "RISK", "RELATED_URL", "REMEDIATION_RECOMMENDATION_TEXT",
    "REMEDIATION_RECOMMENDATION_URL", "REMEDIATION_CODE_NATIVEIAC",
    "REMEDIATION_CODE_TERRAFORM", "REMEDIATION_CODE_CLI", "REMEDIATION_CODE_OTHER",
    "COMPLIANCE", "CATEGORIES", "DEPENDS_ON", "RELATED_TO", "NOTES",
    "PROWLER_VERSION",
]

TS = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
PROWLER_VER = "4.3.1"

# Two synthetic accounts
ACCT_A_UID   = "123456789012"
ACCT_A_NAME  = "acme-prod"
ACCT_A_EMAIL = "aws-root@acme.example"

ACCT_B_UID   = "987654321098"
ACCT_B_NAME  = "acme-dev"
ACCT_B_EMAIL = "aws-dev@acme.example"

ORG_UID   = "o-exampleorg123"
ORG_NAME  = "acme-org"


def row(
    account_uid=ACCT_A_UID,
    account_name=ACCT_A_NAME,
    account_email=ACCT_A_EMAIL,
    check_id="s3_bucket_public_access",
    check_title="Ensure S3 bucket does not allow public access",
    check_type="Software and Configuration Checks",
    status="FAIL",
    status_extended="Bucket my-bucket is publicly accessible.",
    muted="False",
    service_name="s3",
    subservice_name=None,
    severity="high",
    resource_type="AWSS3Bucket",
    resource_uid=None,
    resource_name=None,
    resource_details=None,
    resource_tags="Environment:prod | Owner:platform-team",
    partition="aws",
    region="ap-southeast-2",
    description="This control checks whether Amazon S3 buckets have public access blocked.",
    risk="If this setting is enabled, any policies or access control lists (ACLs) "
         "that grant public access to this bucket and the objects that it contains are blocked.",
    related_url="https://docs.aws.amazon.com/AmazonS3/latest/userguide/access-control-block-public-access.html",
    remediation_text="Enable S3 Block Public Access at the bucket level.",
    remediation_url="https://docs.aws.amazon.com/AmazonS3/latest/userguide/access-control-block-public-access.html",
    remediation_nativeiac=None,
    remediation_terraform=None,
    remediation_cli="aws s3api put-public-access-block --bucket <bucket-name> --public-access-block-configuration "
                    "BlockPublicAcls=true,IgnorePublicAcls=true,BlockPublicPolicy=true,RestrictPublicBuckets=true",
    remediation_other=None,
    compliance="CIS-2.0 | 2.1.5 | NIST-800-53 | SC-7",
    categories="internet-exposed | encryption",
    depends_on=None,
    related_to=None,
    notes=None,
    account_tags="Terraform:true | Created_By:platform",
    org_uid=ORG_UID,
    org_name=ORG_NAME,
    auth_method="profile:None",
) -> dict:
    """Build a single row dict. ARN auto-generated if resource_uid and resource_name both None."""
    if resource_uid is None and resource_name is None:
        resource_name = f"synthetic-resource-{check_id[:10]}"

    finding_uid = (
        f"prowler-aws-{check_id}-{account_uid}-{region or 'global'}-"
        f"{(resource_uid or resource_name or 'nores')[:20]}"
    )

    return {
        "AUTH_METHOD":                      auth_method,
        "TIMESTAMP":                        TS,
        "ACCOUNT_UID":                      account_uid,
        "ACCOUNT_NAME":                     account_name,
        "ACCOUNT_EMAIL":                    account_email,
        "ACCOUNT_ORGANIZATION_UID":         org_uid,
        "ACCOUNT_ORGANIZATION_NAME":        org_name,
        "ACCOUNT_TAGS":                     account_tags,
        "FINDING_UID":                      finding_uid,
        "PROVIDER":                         "aws",
        "CHECK_ID":                         check_id,
        "CHECK_TITLE":                      check_title,
        "CHECK_TYPE":                       check_type,
        "STATUS":                           status,
        "STATUS_EXTENDED":                  status_extended,
        "MUTED":                            muted,
        "SERVICE_NAME":                     service_name,
        "SUBSERVICE_NAME":                  subservice_name,
        "SEVERITY":                         severity,
        "RESOURCE_TYPE":                    resource_type,
        "RESOURCE_UID":                     resource_uid,
        "RESOURCE_NAME":                    resource_name,
        "RESOURCE_DETAILS":                 resource_details,
        "RESOURCE_TAGS":                    resource_tags,
        "PARTITION":                        partition,
        "REGION":                           region,
        "DESCRIPTION":                      description,
        "RISK":                             risk,
        "RELATED_URL":                      related_url,
        "REMEDIATION_RECOMMENDATION_TEXT":  remediation_text,
        "REMEDIATION_RECOMMENDATION_URL":   remediation_url,
        "REMEDIATION_CODE_NATIVEIAC":       remediation_nativeiac,
        "REMEDIATION_CODE_TERRAFORM":       remediation_terraform,
        "REMEDIATION_CODE_CLI":             remediation_cli,
        "REMEDIATION_CODE_OTHER":           remediation_other,
        "COMPLIANCE":                       compliance,
        "CATEGORIES":                       categories,
        "DEPENDS_ON":                       depends_on,
        "RELATED_TO":                       related_to,
        "NOTES":                            notes,
        "PROWLER_VERSION":                  PROWLER_VER,
    }


def build_rows() -> list[dict]:
    """Build all test rows covering every code path."""
    rows = []

    # ── S3 checks ────────────────────────────────────────────────────

    # 1. Normal FAIL — S3 public access — bucket with ARN — Account A
    rows.append(row(
        check_id="s3_bucket_public_access",
        check_title="Ensure S3 bucket does not allow public access",
        severity="high",
        status="FAIL",
        status_extended="Bucket acme-logs is publicly accessible.",
        resource_uid=f"arn:aws:s3:::acme-logs",
        resource_name="acme-logs",
        resource_type="AWSS3Bucket",
        region="ap-southeast-2",
    ))

    # 2. Same check, different bucket — same account (will group into same output row)
    rows.append(row(
        check_id="s3_bucket_public_access",
        check_title="Ensure S3 bucket does not allow public access",
        severity="high",
        status="FAIL",
        status_extended="Bucket acme-backups is publicly accessible.",
        resource_uid=f"arn:aws:s3:::acme-backups",
        resource_name="acme-backups",
        resource_type="AWSS3Bucket",
        region="ap-southeast-2",
    ))

    # 3. Same check, Account B (multi-account — different dedup key)
    rows.append(row(
        account_uid=ACCT_B_UID,
        account_name=ACCT_B_NAME,
        account_email=ACCT_B_EMAIL,
        check_id="s3_bucket_public_access",
        check_title="Ensure S3 bucket does not allow public access",
        severity="high",
        status="FAIL",
        status_extended="Bucket dev-uploads is publicly accessible.",
        resource_uid=f"arn:aws:s3:::dev-uploads",
        resource_name="dev-uploads",
        resource_type="AWSS3Bucket",
        region="ap-southeast-2",
    ))

    # 4. S3 encryption — MUTED(FAIL) — should be included in working set
    rows.append(row(
        check_id="s3_bucket_default_encryption",
        check_title="Ensure S3 bucket has default encryption enabled",
        severity="medium",
        status="MUTED(FAIL)",
        status_extended="Bucket acme-archive has no default encryption.",
        muted="True",
        resource_uid="arn:aws:s3:::acme-archive",
        resource_name="acme-archive",
        resource_type="AWSS3Bucket",
        region="ap-southeast-2",
        categories="encryption",
        compliance="CIS-2.0 | 2.1.1",
    ))

    # 5. S3 versioning — PASS — should be EXCLUDED downstream
    rows.append(row(
        check_id="s3_bucket_versioning_enabled",
        check_title="Ensure S3 bucket has versioning enabled",
        severity="low",
        status="PASS",
        status_extended="Bucket acme-logs has versioning enabled.",
        resource_uid="arn:aws:s3:::acme-logs",
        resource_name="acme-logs",
        resource_type="AWSS3Bucket",
        region="ap-southeast-2",
        categories="resilience",
    ))

    # ── IAM checks (global — blank REGION) ───────────────────────────

    # 6. IAM MFA — FAIL — global check — REGION intentionally blank
    rows.append(row(
        check_id="iam_user_mfa_enabled_console_access",
        check_title="Ensure MFA is enabled for all IAM users with console access",
        check_type="Software and Configuration Checks",
        service_name="iam",
        severity="critical",
        status="FAIL",
        status_extended="IAM user john.smith does not have MFA enabled.",
        resource_type="AWSIAMUser",
        resource_uid="arn:aws:iam::123456789012:user/john.smith",
        resource_name="john.smith",
        region=None,        # ← structural blank — IAM is global
        categories="identity-management",
        compliance="CIS-2.0 | 1.10 | NIST-800-53 | IA-2",
    ))

    # 7. IAM root account — FAIL — no resource ID at all (account-level singleton)
    r7 = row(
        check_id="iam_root_mfa_enabled",
        check_title="Ensure MFA is enabled for the root account",
        service_name="iam",
        severity="critical",
        status="FAIL",
        status_extended="Root account does not have MFA enabled.",
        resource_type="AWSIAMAccountSummary",
        resource_uid=None,  # ← no ARN for account-level check
        resource_name=None, # ← no name either (account singleton)
        region=None,
        categories="identity-management",
        compliance="CIS-2.0 | 1.5",
    )
    # Override: force both resource fields to None (row() auto-generates a name otherwise)
    r7["RESOURCE_UID"] = None
    r7["RESOURCE_NAME"] = None
    rows.append(r7)

    # 8. IAM access key rotation — MUTED=True, STATUS=FAIL (reconciliation test)
    rows.append(row(
        check_id="iam_user_access_key_age_90_days",
        check_title="Ensure IAM user access keys are rotated every 90 days",
        service_name="iam",
        severity="medium",
        status="FAIL",      # ← STATUS says FAIL
        muted="True",       # ← but MUTED=True — should reconcile to MUTED(FAIL)
        status_extended="Access key AKIAIOSFODNN7EXAMPLE for user jane.doe is 120 days old.",
        resource_type="AWSIAMUser",
        resource_uid="arn:aws:iam::123456789012:user/jane.doe",
        resource_name="jane.doe",
        region=None,
        categories="identity-management",
    ))

    # ── EC2 checks ───────────────────────────────────────────────────

    # 9. EC2 security group — FAIL — with ARN
    rows.append(row(
        check_id="ec2_securitygroup_allow_ingress_from_internet_to_tcp_port_22",
        check_title="Ensure no security groups allow ingress from 0.0.0.0/0 to port 22",
        service_name="ec2",
        severity="critical",
        status="FAIL",
        status_extended="Security group sg-0abc1234 allows unrestricted SSH access.",
        resource_type="AWSEC2SecurityGroup",
        resource_uid="arn:aws:ec2:ap-southeast-2:123456789012:security-group/sg-0abc1234",
        resource_name="sg-0abc1234",
        region="ap-southeast-2",
        categories="internet-exposed | network-security",
        compliance="CIS-2.0 | 5.2 | NIST-800-53 | SC-7",
    ))

    # 10. EC2 instance — FAIL — ARN-less (only resource name — ARN fallback test)
    rows.append(row(
        check_id="ec2_instance_imdsv2_enabled",
        check_title="Ensure EC2 instances use IMDSv2",
        service_name="ec2",
        severity="high",
        status="FAIL",
        status_extended="Instance i-0abc1234 does not enforce IMDSv2.",
        resource_type="AWSEC2Instance",
        resource_uid=None,          # ← no ARN
        resource_name="i-0abc1234", # ← only name (ARN fallback test)
        region="ap-southeast-2",
        categories="configuration",
    ))

    # ── RDS checks ───────────────────────────────────────────────────

    # 11. RDS encryption — FAIL — blank DESCRIPTION (Category 2 data quality blank)
    rows.append(row(
        check_id="rds_instance_storage_encrypted",
        check_title="Ensure RDS instances are encrypted at rest",
        service_name="rds",
        severity="high",
        status="FAIL",
        status_extended="RDS instance prod-db is not encrypted at rest.",
        resource_type="AWSRDSDBInstance",
        resource_uid="arn:aws:rds:ap-southeast-2:123456789012:db:prod-db",
        resource_name="prod-db",
        region="ap-southeast-2",
        description=None,           # ← Category 2: data quality blank
        risk=None,                  # ← Category 2: data quality blank
        categories="encryption",
        compliance="CIS-2.0 | 2.3.1",
    ))

    # 12. RDS public — FAIL — duplicate of a later row (dedup test)
    rows.append(row(
        check_id="rds_instance_no_public_access",
        check_title="Ensure RDS instances are not publicly accessible",
        service_name="rds",
        severity="critical",
        status="FAIL",
        status_extended="RDS instance prod-db is publicly accessible.",
        resource_type="AWSRDSDBInstance",
        resource_uid="arn:aws:rds:ap-southeast-2:123456789012:db:prod-db",
        resource_name="prod-db",
        region="ap-southeast-2",
        categories="internet-exposed",
    ))

    # 13. Same as row 12 — exact duplicate (same check_id + resource — dedup test)
    rows.append(row(
        check_id="rds_instance_no_public_access",
        check_title="Ensure RDS instances are not publicly accessible",
        service_name="rds",
        severity="critical",
        status="FAIL",
        status_extended="RDS instance prod-db is publicly accessible.",
        resource_type="AWSRDSDBInstance",
        resource_uid="arn:aws:rds:ap-southeast-2:123456789012:db:prod-db",
        resource_name="prod-db",
        region="ap-southeast-2",
        categories="internet-exposed",
    ))

    # ── VPC checks ───────────────────────────────────────────────────

    # 14. VPC flow logs — FAIL — internet-exposed category (likelihood override test)
    rows.append(row(
        check_id="vpc_flow_logs_enabled",
        check_title="Ensure VPC flow logging is enabled in all VPCs",
        service_name="vpc",
        severity="medium",
        status="FAIL",
        status_extended="VPC vpc-0abc1234 does not have flow logging enabled.",
        resource_type="AWSVPC",
        resource_uid="arn:aws:ec2:ap-southeast-2:123456789012:vpc/vpc-0abc1234",
        resource_name="vpc-0abc1234",
        region="ap-southeast-2",
        categories="logging | internet-exposed",
        compliance="CIS-2.0 | 3.9",
    ))

    # ── Edge cases ────────────────────────────────────────────────────

    # 15. Formula injection in CHECK_TITLE (renderer safety test)
    rows.append(row(
        check_id="cloudtrail_multi_region_enabled",
        check_title="=SUM(1,2) CloudTrail multi-region logging check",  # injection test
        service_name="cloudtrail",
        severity="high",
        status="FAIL",
        status_extended="CloudTrail is not enabled in all regions.",
        resource_type="AWSCloudTrailTrail",
        resource_uid="arn:aws:cloudtrail:ap-southeast-2:123456789012:trail/management-events",
        resource_name="management-events",
        region="ap-southeast-2",
        categories="logging",
        compliance="CIS-2.0 | 3.1",
    ))

    # 16. Long DESCRIPTION (near 32K chars — truncation test)
    long_desc = (
        "This control checks whether AWS CloudTrail is configured to log management "
        "events across all regions. " * 500  # ~28K chars
    )
    rows.append(row(
        check_id="cloudtrail_log_file_validation_enabled",
        check_title="Ensure CloudTrail log file validation is enabled",
        service_name="cloudtrail",
        severity="low",
        status="FAIL",
        status_extended="Trail management-events does not have log file validation enabled.",
        resource_type="AWSCloudTrailTrail",
        resource_uid="arn:aws:cloudtrail:ap-southeast-2:123456789012:trail/management-events",
        resource_name="management-events",
        region="ap-southeast-2",
        description=long_desc,
        categories="logging",
    ))

    # 17. Account-level check — no region, no resource (Config service)
    rows.append(row(
        check_id="config_recorder_all_regions_enabled",
        check_title="Ensure AWS Config is enabled in all regions",
        service_name="account",
        severity="medium",
        status="FAIL",
        status_extended="AWS Config recorder is not enabled in all regions.",
        resource_type="AWSAccount",
        resource_uid=None,
        resource_name=None,
        region=None,
        categories="configuration",
        compliance="CIS-2.0 | 3.5",
    ))

    # 18. MANUAL status (should be included, flagged for review)
    rows.append(row(
        check_id="guardduty_is_enabled",
        check_title="Ensure GuardDuty is enabled",
        service_name="guardduty",
        severity="high",
        status="MANUAL",
        status_extended="Manually verify GuardDuty is enabled in all regions.",
        resource_type="AWSGuardDutyDetector",
        resource_uid=None,
        resource_name="guardduty-detector",
        region="ap-southeast-2",
        description=None,  # Category 2 blank
        categories="threat-detection",
    ))

    # 19. Unknown future column (extra_fields test)
    r = row(
        check_id="securityhub_enabled",
        check_title="Ensure AWS Security Hub is enabled",
        service_name="securityhub",
        severity="medium",
        status="FAIL",
        status_extended="Security Hub is not enabled.",
        resource_uid=None,
        resource_name="securityhub",
        region="ap-southeast-2",
        categories="configuration",
    )
    r["FUTURE_COLUMN_V5"] = "some-future-value"  # unknown column test
    rows.append(r)

    # 20. Informational severity (likelihood → Low)
    rows.append(row(
        check_id="s3_bucket_object_lock",
        check_title="Ensure S3 bucket has Object Lock configured",
        severity="informational",
        status="FAIL",
        status_extended="Bucket acme-logs does not have Object Lock enabled.",
        resource_uid="arn:aws:s3:::acme-logs",
        resource_name="acme-logs",
        resource_type="AWSS3Bucket",
        region="ap-southeast-2",
        categories="resilience",
    ))

    return rows


def generate(output_path: str | Path) -> Path:
    """
    Write the synthetic dataset to an XLSX file.
    Returns the output path.
    """
    output_path = Path(output_path)
    rows_data = build_rows()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "prowler-output"

    # Header row — bold
    header_font = Font(bold=True)

    # Determine all columns (including any unknown extras)
    all_cols = list(HEADERS)
    extra_cols = []
    for r in rows_data:
        for k in r:
            if k not in all_cols and k not in extra_cols:
                extra_cols.append(k)
    all_cols.extend(extra_cols)

    for col_idx, col_name in enumerate(all_cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font

    # Data rows
    for row_idx, row_dict in enumerate(rows_data, 2):
        for col_idx, col_name in enumerate(all_cols, 1):
            value = row_dict.get(col_name)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            # Force formula-like strings to be stored as text, not evaluated as formulas
            if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
                cell.data_type = "s"

    wb.save(str(output_path))
    print(f"Synthetic dataset written: {output_path}")
    print(f"  Rows: {len(rows_data)}")
    print(f"  Columns: {len(all_cols)} ({len(extra_cols)} unknown extras)")

    # Print summary
    status_counts: dict[str, int] = {}
    for r in rows_data:
        s = r.get("STATUS", "unknown")
        status_counts[s] = status_counts.get(s, 0) + 1
    print("  Status breakdown:")
    for s, c in sorted(status_counts.items()):
        print(f"    {s}: {c}")

    return output_path


if __name__ == "__main__":
    out = Path(__file__).parent.parent / "data" / "synthetic" / "synthetic_prowler.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    generate(out)
