"""
stage1_ingest.py — Stage 1: Ingest & Parse

Responsibilities:
    - Accept one or more Prowler CSV/XLSX files
    - Compute SHA-256 hash of each file (immutable evidence anchor)
    - Detect the data sheet by content (never by name or position)
    - Parse all columns BY NAME, never by position
    - Reconcile MUTED vs STATUS fields (MUTED=True always wins)
    - Classify blank values (Category 1/2/3)
    - Build stable_finding_key and dedup_key
    - Emit a list of CanonicalFinding objects with raw fields locked

Contract:
    ingest(file_path: str | Path, run_id: str) -> IngestResult

Nothing downstream of this function may modify raw_ fields.
"""

from __future__ import annotations

import csv
import io
import hashlib
import logging
import re
import uuid
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl
import pandas as pd

from models import (
    AuditEvent,
    BlankCategory,
    CanonicalFinding,
    ReportInclusion,
    ScannerStatus,
)

logger = logging.getLogger(__name__)

# ── Constants ─────────────────────────────────────────────────────────

# The exact 41 column names Prowler uses (as of v4.x).
# Lookup is always by name — position is irrelevant.
PROWLER_COLUMNS: set[str] = {
    "AUTH_METHOD", "TIMESTAMP", "ACCOUNT_UID", "ACCOUNT_NAME",
    "ACCOUNT_EMAIL", "ACCOUNT_ORGANIZATION_UID", "ACCOUNT_ORGANIZATION_NAME",
    "ACCOUNT_TAGS", "FINDING_UID", "PROVIDER", "CHECK_ID", "CHECK_TITLE",
    "CHECK_TYPE", "STATUS", "STATUS_EXTENDED", "MUTED", "SERVICE_NAME",
    "SUBSERVICE_NAME", "SEVERITY", "RESOURCE_TYPE", "RESOURCE_UID",
    "RESOURCE_NAME", "RESOURCE_DETAILS", "RESOURCE_TAGS", "PARTITION",
    "REGION", "DESCRIPTION", "RISK", "RELATED_URL",
    "REMEDIATION_RECOMMENDATION_TEXT", "REMEDIATION_RECOMMENDATION_URL",
    "REMEDIATION_CODE_NATIVEIAC", "REMEDIATION_CODE_TERRAFORM",
    "REMEDIATION_CODE_CLI", "REMEDIATION_CODE_OTHER", "COMPLIANCE",
    "CATEGORIES", "DEPENDS_ON", "RELATED_TO", "NOTES", "PROWLER_VERSION",
}

# Fields that are structurally blank for certain check types (Category 1).
# These are NEVER flagged as data quality issues.
STRUCTURALLY_BLANK_FIELDS: set[str] = {
    "ACCOUNT_ORGANIZATION_UID",   # blank if no AWS Org
    "ACCOUNT_ORGANIZATION_NAME",  # blank if no AWS Org
    "SUBSERVICE_NAME",            # blank for most check types
    "RESOURCE_DETAILS",           # structurally optional
    "DEPENDS_ON",                 # rarely populated
    "RELATED_TO",                 # rarely populated
    "NOTES",                      # analyst field, always blank from scanner
}

# Fields that are Category 2 (data quality blank — should have a value).
DATA_QUALITY_FIELDS: set[str] = {
    "DESCRIPTION",
    "RISK",
    "REMEDIATION_RECOMMENDATION_TEXT",
}

# Services that produce global/IAM findings (no region).
GLOBAL_SERVICES: set[str] = {
    "iam", "account", "organizations", "budgets", "cost-optimization",
    "cloudfront", "route53", "waf",
}

# ── Prowler version compatibility: old column names → new ───────────
# Maps Prowler v3.x column names to their v4.x equivalents.
# Applied during CSV/JSON parsing before the main column map.
COLUMN_ALIASES: dict[str, str] = {
    # v3 → v4 renames
    "ASSESSMENT_START_TIME":                    "TIMESTAMP",
    "FINDING_UNIQUE_ID":                        "FINDING_UID",
    "ACCOUNT_ID":                               "ACCOUNT_UID",
    "ACCOUNT_ORG":                              "ACCOUNT_ORGANIZATION_UID",
    "RESOURCE_ARN":                             "RESOURCE_UID",
    "RESOURCE_ID":                              "RESOURCE_NAME",
    "PROFILE":                                  "AUTH_METHOD",
    "ACCOUNT_ARN":                              "ACCOUNT_ORGANIZATION_UID",  # closest mapping
    "REMEDIATION_RECOMMENDATION_CODE_NATIVEIAC":"REMEDIATION_CODE_NATIVEIAC",
    "REMEDIATION_RECOMMENDATION_CODE_TERRAFORM":"REMEDIATION_CODE_TERRAFORM",
    "REMEDIATION_RECOMMENDATION_CODE_CLI":      "REMEDIATION_CODE_CLI",
    "REMEDIATION_RECOMMENDATION_CODE_OTHER":    "REMEDIATION_CODE_OTHER",
}

def _apply_column_aliases(row: dict) -> dict:
    """
    Rename old Prowler v3 column names to their v4 equivalents.
    Called once per row before the main column mapping.
    Original key is dropped; new key is added only if not already present.
    """
    result = {}
    for k, v in row.items():
        new_key = COLUMN_ALIASES.get(k.strip().upper(), k)
        # If both old and new keys exist, new key wins
        if new_key not in result:
            result[new_key] = v
    return result


# ARN pattern for resource identity normalisation.

ARN_PATTERN = re.compile(r"^arn:[a-z0-9\-]+:[a-z0-9\-]+:[^:]*:[^:]*:.+$", re.IGNORECASE)


# ── Result types ─────────────────────────────────────────────────────

@dataclass
class IngestWarning:
    code: str
    message: str
    row: Optional[int] = None
    column: Optional[str] = None


@dataclass
class IngestResult:
    run_id:             str
    source_file:        str
    source_file_hash:   str
    scanner:            str
    scanner_version:    str
    sheet_name:         str
    total_rows_read:    int
    findings:           list[CanonicalFinding]
    warnings:           list[IngestWarning]
    ingested_at:        datetime = field(default_factory=datetime.utcnow)
    unknown_columns:    list[str] = field(default_factory=list)

    @property
    def finding_count(self) -> int:
        return len(self.findings)


# ── File hash ─────────────────────────────────────────────────────────

def _sha256(path: Path) -> str:
    """Compute SHA-256 of the file. This is the root-of-trust for the entire run."""
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


# ── Sheet detection ───────────────────────────────────────────────────

def _detect_data_sheet(wb: openpyxl.Workbook) -> tuple[str, openpyxl.worksheet.worksheet.Worksheet]:
    """
    Detect the sheet containing Prowler data.

    Strategy:
        1. Look for a sheet whose first non-empty row contains >= 10
           known Prowler column names.
        2. Prefer sheets with more data rows if multiple candidates.
        3. Raise ValueError with a clear message if ambiguous or not found.

    We never assume sheet name or position.
    """
    candidates: list[tuple[str, openpyxl.worksheet.worksheet.Worksheet, int, int]] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Find first non-empty row
        header_row = None
        for row in ws.iter_rows(values_only=True):
            if any(v is not None for v in row):
                header_row = [str(v).strip() if v is not None else "" for v in row]
                break
        if header_row is None:
            continue
        # Count how many cells match known Prowler columns
        matches = sum(1 for v in header_row if v.upper() in PROWLER_COLUMNS)
        if matches >= 10:  # threshold: at least 10 of 41 known columns
            # Count data rows
            data_rows = sum(
                1 for row in ws.iter_rows(min_row=2, values_only=True)
                if any(v is not None for v in row)
            )
            candidates.append((sheet_name, ws, matches, data_rows))
            logger.debug(
                "Sheet '%s': %d column matches, %d data rows",
                sheet_name, matches, data_rows,
            )

    if not candidates:
        raise ValueError(
            "No sheet found with Prowler column headers. "
            "Expected at least 10 of the known 41 Prowler column names. "
            "Please verify this is a valid Prowler output file."
        )

    if len(candidates) > 1:
        # Pick the one with the most data rows; flag if tied
        candidates.sort(key=lambda x: (x[2], x[3]), reverse=True)
        best = candidates[0]
        second = candidates[1]
        if best[3] == second[3]:
            logger.warning(
                "Multiple sheets match Prowler format with equal row counts: %s. "
                "Using '%s'. Add sheet_hint to config if this is wrong.",
                [c[0] for c in candidates],
                best[0],
            )

    name, ws, _, _ = candidates[0]
    logger.info("Data sheet detected: '%s'", name)
    return name, ws


# ── Raw value helpers ─────────────────────────────────────────────────

def _clean(value: Any) -> Optional[str]:
    """Normalise a raw cell value to a stripped string or None."""
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def _parse_bool(value: Any) -> Optional[bool]:
    """Parse True/False from various representations Prowler uses."""
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    s = str(value).strip().lower()
    if s in ("true", "1", "yes"):
        return True
    if s in ("false", "0", "no"):
        return False
    return None


def _parse_pipe_list(value: Optional[str]) -> list[str]:
    """Parse a pipe-delimited list, tolerating extra spaces."""
    if not value:
        return []
    return [item.strip() for item in value.split("|") if item.strip()]


def _parse_tags(value: Optional[str]) -> dict[str, str]:
    """
    Parse pipe-delimited key:value tag strings.
    Handles: nested colons in values (e.g. ARNs), missing values, spaces.
    """
    result: dict[str, str] = {}
    if not value:
        return result
    for item in value.split("|"):
        item = item.strip()
        if not item:
            continue
        if ":" in item:
            key, _, val = item.partition(":")
            result[key.strip()] = val.strip()
        else:
            result[item] = ""
    return result


def _parse_compliance(value: Optional[str]) -> tuple[list[str], bool]:
    """
    Attempt to parse the COMPLIANCE field.
    Returns (parsed_list, success_flag).
    Stores raw string regardless — parsing failure never blocks the pipeline.
    """
    if not value:
        return [], False
    try:
        # Try pipe-delimited first
        items = _parse_pipe_list(value)
        if items:
            return items, True
        # Try JSON-like
        import json
        parsed = json.loads(value)
        if isinstance(parsed, list):
            return [str(i) for i in parsed], True
        if isinstance(parsed, dict):
            return [f"{k}:{v}" for k, v in parsed.items()], True
    except Exception:
        pass
    # Fallback: return as single item
    return [value.strip()], False


# ── MUTED / STATUS reconciliation ────────────────────────────────────

def _reconcile_status(
    raw_status: Optional[str],
    raw_muted: Optional[str],
    row_num: int,
    warnings: list[IngestWarning],
) -> tuple[ScannerStatus, bool]:
    """
    Reconcile STATUS and MUTED columns into a single ScannerStatus.

    Rule: MUTED=True always wins, regardless of STATUS value.

    Returns (scanner_status, muted_reconciled_flag).
    muted_reconciled_flag is True if MUTED=True overrode a non-MUTED STATUS.
    """
    muted_bool = _parse_bool(raw_muted)
    status_upper = (raw_status or "").strip().upper()

    # Determine original status from the STATUS column
    if status_upper.startswith("MUTED(") and status_upper.endswith(")"):
        inner = status_upper[6:-1]
        status_from_col = f"MUTED({inner})"
    elif status_upper in ("FAIL", "PASS", "MANUAL"):
        status_from_col = status_upper
    else:
        status_from_col = "UNKNOWN"

    # Apply MUTED=True override
    if muted_bool is True:
        # Determine underlying status
        if status_upper.startswith("MUTED(") and status_upper.endswith(")"):
            underlying = status_upper[6:-1]
        else:
            underlying = status_upper if status_upper in ("FAIL", "PASS", "MANUAL") else "FAIL"

        reconciled = f"MUTED({underlying})"
        was_overridden = not status_upper.startswith("MUTED(")

        if was_overridden:
            warnings.append(IngestWarning(
                code="MUTED_STATUS_RECONCILED",
                message=(
                    f"MUTED=True but STATUS='{raw_status}' (not prefixed MUTED). "
                    f"Reconciled to '{reconciled}'. Original STATUS preserved in raw_status."
                ),
                row=row_num,
            ))

        try:
            return ScannerStatus(reconciled), was_overridden
        except ValueError:
            return ScannerStatus.MUTED_FAIL, was_overridden

    # MUTED=False or None — use STATUS column as-is
    try:
        return ScannerStatus(status_from_col), False
    except ValueError:
        if status_upper:
            warnings.append(IngestWarning(
                code="UNKNOWN_STATUS",
                message=f"Unrecognised STATUS value '{raw_status}'. Set to UNKNOWN.",
                row=row_num,
            ))
        return ScannerStatus.UNKNOWN, False


# ── Blank value classification ─────────────────────────────────────────

def _classify_blank(
    column_name: str,
    value: Optional[str],
    service_name: Optional[str],
    check_type: Optional[str],
) -> BlankCategory:
    """Classify why a field is blank."""
    if value is not None:
        return BlankCategory.POPULATED

    col_upper = column_name.upper()

    if col_upper in STRUCTURALLY_BLANK_FIELDS:
        return BlankCategory.STRUCTURAL

    # REGION is structural blank for global services
    if col_upper == "REGION" and service_name and service_name.lower() in GLOBAL_SERVICES:
        return BlankCategory.STRUCTURAL

    if col_upper in DATA_QUALITY_FIELDS:
        return BlankCategory.DATA_QUALITY

    return BlankCategory.BY_DESIGN


# ── Resource identity normalisation ───────────────────────────────────

def _normalise_resource_id(
    resource_uid: Optional[str],
    resource_name: Optional[str],
) -> tuple[str, bool]:
    """
    Normalise resource identity.
    Prefer ARN (RESOURCE_UID). Fall back to RESOURCE_NAME.
    Returns (normalised_id, arn_fallback_used).
    """
    if resource_uid:
        uid_clean = resource_uid.strip()
        if ARN_PATTERN.match(uid_clean):
            # Valid ARN — normalise: lowercase, strip trailing slashes
            return uid_clean.rstrip("/"), False
        # Not an ARN — still use it as identifier
        if uid_clean:
            return uid_clean, False

    if resource_name and resource_name.strip():
        return resource_name.strip(), True

    return "", True  # neither present


# ── Key construction ───────────────────────────────────────────────────

def _build_stable_finding_key(
    provider: Optional[str],
    account_name: Optional[str],
    service_name: Optional[str],
    check_id: Optional[str],
    normalised_resource_id: str,
) -> str:
    """
    Build a stable_finding_key that is consistent across repeated scans
    of the same environment.

    Format: "provider:account_name:service:check_id:resource_id"

    Uses account_name (not account_uid) so it survives account ID changes.
    Falls back to "unknown" for any missing component.
    """
    parts = [
        (provider or "unknown").lower().strip(),
        (account_name or "unknown").lower().strip().replace(" ", "_"),
        (service_name or "unknown").lower().strip(),
        (check_id or "unknown").lower().strip(),
        (normalised_resource_id or "no_resource").lower().strip(),
    ]
    return ":".join(parts)


def _build_dedup_key(
    account_uid: Optional[str],
    check_id: Optional[str],
    normalised_resource_id: str,
    region_normalised: str,
    service_name: Optional[str],
) -> str:
    """
    Build a dedup_key for within-run collision detection.

    Cases:
        AWS resource:       account_uid + check_id + resource_id + region
        IAM/global:         account_uid + check_id + resource_id + "global"
        Account singleton:  account_uid + check_id  (no resource)
    """
    svc = (service_name or "").lower().strip()
    acct = (account_uid or "unknown").strip()
    chk  = (check_id or "unknown").lower().strip()
    res  = normalised_resource_id.lower().strip()
    rgn  = region_normalised.lower().strip()

    if not res:
        # Account-scoped singleton: no resource ID at all
        return f"{acct}:{chk}"

    return f"{acct}:{chk}:{res}:{rgn}"


# ── Main parser ───────────────────────────────────────────────────────

_COLUMN_MAP = {
    # Maps Prowler column name (upper) → CanonicalFinding raw_ field name
    "AUTH_METHOD":                      "raw_auth_method",
    "TIMESTAMP":                        "raw_timestamp",
    "ACCOUNT_UID":                      "raw_account_uid",
    "ACCOUNT_NAME":                     "raw_account_name",
    "ACCOUNT_EMAIL":                    "raw_account_email",
    "ACCOUNT_ORGANIZATION_UID":         "raw_account_organization_uid",
    "ACCOUNT_ORGANIZATION_NAME":        "raw_account_organization_name",
    "ACCOUNT_TAGS":                     "raw_account_tags",
    "FINDING_UID":                      "raw_finding_uid",
    "PROVIDER":                         "raw_provider",
    "CHECK_ID":                         "raw_check_id",
    "CHECK_TITLE":                      "raw_check_title",
    "CHECK_TYPE":                       "raw_check_type",
    "STATUS":                           "raw_status",
    "STATUS_EXTENDED":                  "raw_status_extended",
    "MUTED":                            "raw_muted",
    "SERVICE_NAME":                     "raw_service_name",
    "SUBSERVICE_NAME":                  "raw_subservice_name",
    "SEVERITY":                         "raw_severity",
    "RESOURCE_TYPE":                    "raw_resource_type",
    "RESOURCE_UID":                     "raw_resource_uid",
    "RESOURCE_NAME":                    "raw_resource_name",
    "RESOURCE_DETAILS":                 "raw_resource_details",
    "RESOURCE_TAGS":                    "raw_resource_tags",
    "PARTITION":                        "raw_partition",
    "REGION":                           "raw_region",
    "DESCRIPTION":                      "raw_description",
    "RISK":                             "raw_risk",
    "RELATED_URL":                      "raw_related_url",
    "REMEDIATION_RECOMMENDATION_TEXT":  "raw_remediation_recommendation_text",
    "REMEDIATION_RECOMMENDATION_URL":   "raw_remediation_recommendation_url",
    "REMEDIATION_CODE_NATIVEIAC":       "raw_remediation_code_nativeiac",
    "REMEDIATION_CODE_TERRAFORM":       "raw_remediation_code_terraform",
    "REMEDIATION_CODE_CLI":             "raw_remediation_code_cli",
    "REMEDIATION_CODE_OTHER":           "raw_remediation_code_other",
    "COMPLIANCE":                       "raw_compliance",
    "CATEGORIES":                       "raw_categories",
    "DEPENDS_ON":                       "raw_depends_on",
    "RELATED_TO":                       "raw_related_to",
    "NOTES":                            "raw_notes",
    "PROWLER_VERSION":                  "raw_prowler_version",
}


def _parse_rows(
    rows: list[dict[str, Any]],
    source_file: str,
    source_file_hash: str,
    run_id: str,
    sheet_name: str,
    unknown_columns: list[str],
    warnings: list[IngestWarning],
) -> list[CanonicalFinding]:
    """Parse a list of row dicts (keyed by column name) into CanonicalFinding objects."""
    findings: list[CanonicalFinding] = []

    for row_idx, row in enumerate(rows, start=2):  # start=2 because row 1 is header
        # Skip entirely blank rows
        if all(v is None or str(v).strip() == "" for v in row.values()):
            continue

        finding = CanonicalFinding(
            run_id=run_id,
            source_file=source_file,
            source_file_hash=source_file_hash,
            source_row_id=f"Sheet:{sheet_name} Row:{row_idx}",
        )

        # ── Map known columns to raw_ fields ──
        # Apply v3→v4 column aliases first (covers XLSX from old Prowler)
        row = _apply_column_aliases({k: v for k, v in row.items()})
        extra: dict[str, Any] = {}
        for col_name, value in row.items():
            col_upper = col_name.strip().upper()
            cleaned = _clean(value)
            if col_upper in _COLUMN_MAP:
                field_name = _COLUMN_MAP[col_upper]
                setattr(finding, field_name, cleaned)
            elif col_upper:
                extra[col_name] = cleaned
                if col_upper not in {c.upper() for c in unknown_columns}:
                    unknown_columns.append(col_name)

        if extra:
            finding.extra_fields = extra

        # ── Scanner version from PROWLER_VERSION column ──
        if finding.raw_prowler_version:
            finding.scanner_version = finding.raw_prowler_version

        # ── MUTED / STATUS reconciliation ──
        scanner_status, muted_reconciled = _reconcile_status(
            finding.raw_status,
            finding.raw_muted,
            row_idx,
            warnings,
        )
        finding.scanner_status = scanner_status
        finding.muted_reconciled = muted_reconciled

        if muted_reconciled:
            finding.add_audit(
                stage="stage1_ingest",
                field="scanner_status",
                old_value=finding.raw_status,
                new_value=scanner_status.value,
                reason="MUTED=True overrode STATUS column value",
            )

        # ── Blank value classification ──
        finding.blank_description = _classify_blank(
            "DESCRIPTION", finding.raw_description,
            finding.raw_service_name, finding.raw_check_type,
        )
        finding.blank_risk = _classify_blank(
            "RISK", finding.raw_risk,
            finding.raw_service_name, finding.raw_check_type,
        )
        finding.blank_remediation = _classify_blank(
            "REMEDIATION_RECOMMENDATION_TEXT",
            finding.raw_remediation_recommendation_text,
            finding.raw_service_name, finding.raw_check_type,
        )
        finding.blank_region = _classify_blank(
            "REGION", finding.raw_region,
            finding.raw_service_name, finding.raw_check_type,
        )

        # Flag Category 2 blanks for human review
        data_quality_issues = []
        if finding.blank_description == BlankCategory.DATA_QUALITY:
            data_quality_issues.append("DESCRIPTION")
        if finding.blank_risk == BlankCategory.DATA_QUALITY:
            data_quality_issues.append("RISK")
        if finding.blank_remediation == BlankCategory.DATA_QUALITY:
            data_quality_issues.append("REMEDIATION_RECOMMENDATION_TEXT")

        if data_quality_issues:
            finding.flag_for_review(
                reason=f"Category 2 blank fields: {', '.join(data_quality_issues)}",
                stage="stage1_ingest",
            )
            warnings.append(IngestWarning(
                code="DATA_QUALITY_BLANK",
                message=f"Row {row_idx}: blank data quality fields: {data_quality_issues}",
                row=row_idx,
            ))

        # ── Region normalisation ──
        region = finding.raw_region
        svc = (finding.raw_service_name or "").lower().strip()
        if not region or region.strip() == "":
            finding.region_normalised = "global"
        else:
            finding.region_normalised = region.strip()

        # ── Resource ID normalisation ──
        norm_res, arn_fallback = _normalise_resource_id(
            finding.raw_resource_uid,
            finding.raw_resource_name,
        )
        finding.resource_uid_normalised = norm_res
        finding.arn_fallback_used = arn_fallback

        if arn_fallback and (finding.raw_resource_uid or finding.raw_resource_name):
            warnings.append(IngestWarning(
                code="ARN_FALLBACK",
                message=(
                    f"Row {row_idx}: no valid ARN in RESOURCE_UID='{finding.raw_resource_uid}'. "
                    f"Using RESOURCE_NAME='{finding.raw_resource_name}' as resource identifier."
                ),
                row=row_idx,
            ))

        if not norm_res:
            finding.flag_for_review(
                reason="No resource identifier (both RESOURCE_UID and RESOURCE_NAME blank)",
                stage="stage1_ingest",
            )

        # ── Parse multi-value fields ──
        finding.categories_list = _parse_pipe_list(finding.raw_categories)
        finding.compliance_values, finding.compliance_parsed = _parse_compliance(
            finding.raw_compliance
        )
        finding.account_tags_parsed = _parse_tags(finding.raw_account_tags)
        finding.resource_tags_parsed = _parse_tags(finding.raw_resource_tags)

        # ── Build stable_finding_key ──
        finding.stable_finding_key = _build_stable_finding_key(
            provider=finding.raw_provider,
            account_name=finding.raw_account_name,
            service_name=finding.raw_service_name,
            check_id=finding.raw_check_id,
            normalised_resource_id=finding.resource_uid_normalised,
        )

        # ── Build dedup_key ──
        finding.dedup_key = _build_dedup_key(
            account_uid=finding.raw_account_uid,
            check_id=finding.raw_check_id,
            normalised_resource_id=finding.resource_uid_normalised,
            region_normalised=finding.region_normalised,
            service_name=finding.raw_service_name,
        )

        findings.append(finding)

    return findings


# ── XLSX reader ───────────────────────────────────────────────────────

def _read_xlsx(
    path: Path,
    warnings: list[IngestWarning],
) -> tuple[str, list[dict]]:
    """
    Read a Prowler XLSX file.
    Returns (sheet_name, list_of_row_dicts).
    """
    wb = openpyxl.load_workbook(str(path), read_only=False, data_only=True)
    sheet_name, ws = _detect_data_sheet(wb)

    rows_data = list(ws.iter_rows(values_only=True))
    if not rows_data:
        raise ValueError(f"Sheet '{sheet_name}' is empty.")

    # Find header row (first non-empty row)
    header_row_idx = None
    headers = []
    for i, row in enumerate(rows_data):
        if any(v is not None for v in row):
            headers = [str(v).strip() if v is not None else "" for v in row]
            header_row_idx = i
            break

    if header_row_idx is None:
        raise ValueError(f"No header row found in sheet '{sheet_name}'.")

    # Validate that we found Prowler headers
    found_prowler_cols = sum(1 for h in headers if h.upper() in PROWLER_COLUMNS)
    if found_prowler_cols < 10:
        raise ValueError(
            f"Only {found_prowler_cols} Prowler columns found in header. "
            f"Expected at least 10. Headers: {headers[:10]}"
        )

    # Parse data rows
    row_dicts = []
    for row in rows_data[header_row_idx + 1:]:
        row_dict = {}
        for h, v in zip(headers, row):
            if h:  # skip columns with blank headers
                row_dict[h] = v
        # Handle rows wider than header
        if len(row) > len(headers):
            extra_vals = row[len(headers):]
            if any(v is not None for v in extra_vals):
                warnings.append(IngestWarning(
                    code="ROW_WIDER_THAN_HEADER",
                    message=f"Row has {len(row)} values but only {len(headers)} headers. Extra values ignored.",
                ))
        row_dicts.append(row_dict)

    logger.info(
        "XLSX: read %d candidate rows from sheet '%s'",
        len(row_dicts), sheet_name,
    )
    return sheet_name, row_dicts


# ── CSV reader ────────────────────────────────────────────────────────

def _sniff_delimiter(header_line: str) -> str:
    """
    Auto-detect the CSV delimiter from the header line.
    Prowler uses comma by default but semicolon in some locales.
    Checks semicolon, comma, tab in that order by column count.
    """
    candidates = [";", ",", "\t", "|"]
    best_delim = ","
    best_count = 0
    for delim in candidates:
        count = header_line.count(delim)
        if count > best_count:
            best_count = count
            best_delim = delim
    return best_delim


def _fix_line_endings(raw_bytes: bytes) -> str:
    """
    Fix Prowler's known double-CRLF bug (each row ends with \r\r\n).
    Also normalises all other line ending variants to \n.
    Returns clean UTF-8 text.
    """
    # Try UTF-8, fall back to latin-1
    try:
        text = raw_bytes.decode("utf-8")
    except UnicodeDecodeError:
        text = raw_bytes.decode("latin-1")

    # Fix double CRLF: \r\r\n → \n
    text = text.replace("\r\r\n", "\n")
    # Fix remaining CRLF
    text = text.replace("\r\n", "\n")
    # Fix lone CR
    text = text.replace("\r", "\n")
    return text


def _read_csv(
    path: Path,
    warnings: list[IngestWarning],
) -> tuple[str, list[dict]]:
    """
    Read a Prowler CSV file.

    Handles:
        - Semicolon or comma delimiters (auto-detected)
        - Double CRLF line endings (Prowler bug: github.com/prowler-cloud/prowler/issues/*)
        - UTF-8 and latin-1 encoding
        - Prowler v3 column names (mapped to v4 via COLUMN_ALIASES)
    Returns ("csv", list_of_row_dicts).
    """
    with open(path, "rb") as f:
        raw = f.read()

    # Fix line endings — handles the double-CRLF Prowler bug
    text = _fix_line_endings(raw)
    if "\r" in text or "\r\r" in raw.decode("latin-1", errors="replace"):
        warnings.append(IngestWarning(
            code="LINE_ENDING_FIXED",
            message=(
                f"Non-standard line endings detected in {path.name} "
                f"(known Prowler CSV bug). Fixed automatically."
            ),
        ))

    # Remove empty lines
    lines = [l for l in text.split("\n") if l.strip()]
    if not lines:
        raise ValueError("CSV file is empty after cleaning.")

    # Auto-detect delimiter
    delimiter = _sniff_delimiter(lines[0])
    if delimiter != ",":
        warnings.append(IngestWarning(
            code="NON_STANDARD_DELIMITER",
            message=(
                f"CSV delimiter detected as '{delimiter}' (not comma). "
                f"This is normal for some Prowler versions/locales."
            ),
        ))

    clean_text = "\n".join(lines)
    reader = csv.DictReader(io.StringIO(clean_text), delimiter=delimiter)

    if reader.fieldnames is None:
        raise ValueError("CSV file has no header row.")

    # Check for old schema and warn
    headers_upper = {h.strip().upper() for h in reader.fieldnames if h}
    old_schema_hits = headers_upper & set(COLUMN_ALIASES.keys())
    if old_schema_hits:
        warnings.append(IngestWarning(
            code="OLD_PROWLER_SCHEMA",
            message=(
                f"Detected Prowler v3.x column names: {sorted(old_schema_hits)}. "
                f"Automatically mapping to v4.x schema via COLUMN_ALIASES."
            ),
        ))

    # Combined: v3 aliases + v4 columns
    all_known = PROWLER_COLUMNS | set(COLUMN_ALIASES.keys())
    all_known_upper = {c.upper() for c in all_known}
    found_prowler = headers_upper & all_known_upper
    if len(found_prowler) < 10:
        raise ValueError(
            f"Only {len(found_prowler)} recognisable Prowler columns found in CSV. "
            f"Expected at least 10. "
            f"Found: {sorted(headers_upper)[:10]}"
        )

    row_dicts = []
    for row in reader:
        # Strip whitespace from keys; apply column aliases
        cleaned = {k.strip(): v for k, v in row.items() if k}
        aliased = _apply_column_aliases(cleaned)
        row_dicts.append(aliased)

    logger.info(
        "CSV: read %d rows, delimiter='%s', schema=%s",
        len(row_dicts), delimiter,
        "v3-compat" if old_schema_hits else "v4",
    )
    return "csv", row_dicts

# ── OCSF JSON helpers (Prowler v5+) ──────────────────────────────────

_OCSF_UID_RE = re.compile(
    r"^prowler-[a-z]+-(.+?)-\d{10,12}-",
    re.IGNORECASE,
)


def _get(d: dict, *keys, default=None):
    """Safe nested dict getter."""
    cur = d
    for k in keys:
        if not isinstance(cur, dict):
            return default
        cur = cur.get(k, default)
        if cur is None:
            return default
    return cur


def _ocsf_to_flat(item: dict) -> dict:
    """Convert one OCSF finding (Prowler v5 JSON) to flat v4 column names."""
    resources   = item.get("resources") or []
    res0        = resources[0] if resources else {}
    cloud       = item.get("cloud") or {}
    account     = cloud.get("account") or {}
    org         = cloud.get("org") or {}
    finding     = item.get("finding_info") or {}
    remediation = item.get("remediation") or {}
    unmapped    = item.get("unmapped") or {}
    metadata    = item.get("metadata") or {}
    product     = metadata.get("product") or {}

    # CHECK_ID from finding_info.uid
    fuid     = finding.get("uid") or ""
    check_id = ""
    m = _OCSF_UID_RE.match(fuid)
    if m:
        check_id = m.group(1)

    # REGION: resource first, cloud fallback
    region = _get(res0, "region") or cloud.get("region") or None

    # SERVICE_NAME from resource group
    service_name = _get(res0, "group", "name") or None

    # CATEGORIES: list → pipe-delimited
    cats       = unmapped.get("categories") or []
    categories = " | ".join(str(c) for c in cats if c) if cats else None

    # COMPLIANCE: dict → pipe-delimited
    compliance_raw = unmapped.get("compliance") or {}
    if isinstance(compliance_raw, dict):
        parts = []
        for fw, ctrls in compliance_raw.items():
            if isinstance(ctrls, list):
                for c in ctrls:
                    if c:
                        parts.append(f"{fw}: {c}")
            elif ctrls:
                parts.append(f"{fw}: {ctrls}")
        compliance = " | ".join(parts) if parts else None
    else:
        compliance = str(compliance_raw) if compliance_raw else None

    # REMEDIATION URLs
    refs           = remediation.get("references") or []
    additional_urls= unmapped.get("additional_urls") or []
    remediation_url= next((r for r in refs if r), None) or next((u for u in additional_urls if u), None)

    # MUTED: Prowler v5 uses status='Suppressed'
    muted = "True" if (item.get("status") or "").lower() in ("suppressed", "muted") else "False"

    # CHECK_TYPE
    types      = finding.get("types") or []
    check_type = types[0] if types else (item.get("category_name") or None)

    # Labels → pipe-delimited tags
    acc_labels = account.get("labels") or []
    res_labels = res0.get("labels") or []

    # DEPENDS_ON / RELATED_TO
    dep = unmapped.get("depends_on") or []
    rel = unmapped.get("related_to") or []

    return {
        "PROWLER_VERSION":                  product.get("version") or "",
        "AUTH_METHOD":                      None,
        "TIMESTAMP":                        finding.get("created_time_dt") or str(item.get("time_dt") or ""),
        "ACCOUNT_UID":                      account.get("uid") or "",
        "ACCOUNT_NAME":                     account.get("name") or None,
        "ACCOUNT_EMAIL":                    None,
        "ACCOUNT_ORGANIZATION_UID":         org.get("uid") or None,
        "ACCOUNT_ORGANIZATION_NAME":        None,
        "ACCOUNT_TAGS":                     " | ".join(str(l) for l in acc_labels if l) or None,
        "FINDING_UID":                      fuid or None,
        "PROVIDER":                         cloud.get("provider") or unmapped.get("provider") or "aws",
        "CHECK_ID":                         check_id or None,
        "CHECK_TITLE":                      finding.get("title") or None,
        "CHECK_TYPE":                       check_type,
        "STATUS":                           item.get("status_code") or "UNKNOWN",
        "STATUS_EXTENDED":                  item.get("status_detail") or None,
        "MUTED":                            muted,
        "SERVICE_NAME":                     service_name,
        "SUBSERVICE_NAME":                  None,
        "SEVERITY":                         (item.get("severity") or "").lower() or None,
        "RESOURCE_TYPE":                    res0.get("type") or None,
        "RESOURCE_UID":                     res0.get("uid") or None,
        "RESOURCE_NAME":                    res0.get("name") or None,
        "RESOURCE_DETAILS":                 _get(res0, "data", "details") or None,
        "RESOURCE_TAGS":                    " | ".join(str(l) for l in res_labels if l) or None,
        "PARTITION":                        None,
        "REGION":                           region,
        "DESCRIPTION":                      finding.get("desc") or None,
        "RISK":                             item.get("risk_details") or None,
        "RELATED_URL":                      unmapped.get("related_url") or None,
        "REMEDIATION_RECOMMENDATION_TEXT":  remediation.get("desc") or None,
        "REMEDIATION_RECOMMENDATION_URL":   remediation_url,
        "REMEDIATION_CODE_NATIVEIAC":       None,
        "REMEDIATION_CODE_TERRAFORM":       None,
        "REMEDIATION_CODE_CLI":             None,
        "REMEDIATION_CODE_OTHER":           None,
        "COMPLIANCE":                       compliance,
        "CATEGORIES":                       categories,
        "DEPENDS_ON":                       " | ".join(str(d) for d in dep if d) or None,
        "RELATED_TO":                       " | ".join(str(r) for r in rel if r) or None,
        "NOTES":                            unmapped.get("notes") or None,
    }


def _is_ocsf(data: list) -> bool:
    """Detect Prowler v5 OCSF format vs v4 flat JSON."""
    if not data:
        return False
    first = data[0] if isinstance(data, list) else data
    return isinstance(first, dict) and (
        "finding_info" in first or "resources" in first
    )


def _read_json(
    path: Path,
    warnings: list[IngestWarning],
) -> tuple[str, list[dict]]:
    """
    Read a Prowler JSON output file.

    Supports:
        - Prowler v5 OCSF format (finding_info / resources / cloud structure)
        - Prowler v4 flat JSON array (same keys as CSV)
        - Wrapped formats: {"findings": [...]} or {"data": [...]}

    Returns ("json", list_of_row_dicts) where every dict uses v4 column names.
    """
    import json as _json

    with open(path, "rb") as f:
        raw = f.read()

    try:
        text = raw.decode("utf-8")
    except UnicodeDecodeError:
        text = raw.decode("latin-1")
        warnings.append(IngestWarning(
            code="ENCODING_FALLBACK",
            message=f"UTF-8 decode failed for {path.name}. Using latin-1.",
        ))

    try:
        data = _json.loads(text)
    except _json.JSONDecodeError as e:
        raise ValueError(f"JSON parse error in {path.name}: {e}") from e

    # Unwrap container formats
    if isinstance(data, dict):
        if "findings" in data:
            data = data["findings"]
        elif "data" in data:
            data = data["data"]
        else:
            # Single finding wrapped in a dict — treat as one-item list
            data = [data]

    if not isinstance(data, list):
        raise ValueError(
            f"Expected a JSON array in {path.name}, got {type(data).__name__}."
        )

    if not data:
        return "json", []

    # Detect format
    if _is_ocsf(data):
        warnings.append(IngestWarning(
            code="OCSF_FORMAT_DETECTED",
            message=(
                f"Prowler v5 OCSF JSON detected in {path.name}. "
                "Parsing nested structure (finding_info / resources / cloud)."
            ),
        ))
        row_dicts = [_ocsf_to_flat(item) for item in data if isinstance(item, dict)]
        logger.info("OCSF JSON: parsed %d findings", len(row_dicts))
    else:
        # Prowler v4 flat JSON — same keys as CSV
        row_dicts = []
        for item in data:
            if not isinstance(item, dict):
                continue
            flat: dict[str, Any] = {}
            for k, v in item.items():
                if v is None:
                    flat[k] = None
                elif isinstance(v, (dict, list)):
                    import json as _j
                    flat[k] = _j.dumps(v)
                else:
                    flat[k] = str(v)
            row_dicts.append(_apply_column_aliases(flat))

        # Check for v3 column names
        if row_dicts:
            old_hits = set(row_dicts[0].keys()) & set(COLUMN_ALIASES.keys())
            if old_hits:
                warnings.append(IngestWarning(
                    code="OLD_PROWLER_SCHEMA",
                    message=(
                        f"Prowler v3 field names detected in JSON: {sorted(old_hits)}. "
                        "Automatically mapped to v4 schema."
                    ),
                ))
        logger.info("Flat JSON: parsed %d findings", len(row_dicts))

    return "json", row_dicts


# ── Public entry point ────────────────────────────────────────────────

def ingest(
    file_path: str | Path,
    run_id: Optional[str] = None,
    fmt: str = "auto",
) -> IngestResult:
    """
    Stage 1 entry point.

    Args:
        file_path: Path to a Prowler CSV, XLSX, or JSON file.
        run_id:    Optional run ID. Generated if not provided.
        fmt:       Force format: 'json' | 'csv' | 'xlsx' | 'auto' (default).

    Returns:
        IngestResult with all parsed CanonicalFinding objects.

    Supports:
        - Prowler v5 OCSF JSON (nested structure)
        - Prowler v4 flat JSON
        - Prowler v4 XLSX
        - Prowler v4 CSV (comma or semicolon delimiter, double-CRLF fixed)
        - Prowler v3 CSV/JSON (column aliases applied automatically)
    """
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"Input file not found: {path}")

    if run_id is None:
        run_id = str(uuid.uuid4())

    warnings: list[IngestWarning] = []
    unknown_columns: list[str] = []

    # ── Hash the file ──
    logger.info("Computing SHA-256 hash of %s", path.name)
    file_hash = _sha256(path)
    logger.info("SHA-256: %s", file_hash)

    # ── Detect format and read ──
    # fmt overrides auto-detection when explicitly set
    suffix = path.suffix.lower()
    resolved_fmt = fmt if fmt != "auto" else {
        ".xlsx": "xlsx", ".xls": "xlsx",
        ".csv":  "csv",
        ".json": "json",
    }.get(suffix, "auto")

    if resolved_fmt == "json":
        sheet_name, row_dicts = _read_json(path, warnings)
    elif resolved_fmt == "xlsx":
        sheet_name, row_dicts = _read_xlsx(path, warnings)
    elif resolved_fmt == "csv":
        sheet_name, row_dicts = _read_csv(path, warnings)
    else:
        # Unknown extension — try JSON first (most reliable), then XLSX, then CSV
        for reader_fn in (_read_json, _read_xlsx, _read_csv):
            try:
                sheet_name, row_dicts = reader_fn(path, warnings)
                break
            except Exception:
                continue
        else:
            raise ValueError(
                f"Could not parse {path.name} as JSON, XLSX, or CSV. "
                "Use --format json|xlsx|csv to specify the format explicitly."
            )

    # ── Parse rows into CanonicalFinding objects ──
    logger.info("Parsing %d candidate rows", len(row_dicts))
    findings = _parse_rows(
        rows=row_dicts,
        source_file=str(path),
        source_file_hash=file_hash,
        run_id=run_id,
        sheet_name=sheet_name,
        unknown_columns=unknown_columns,
        warnings=warnings,
    )

    # ── Unknown columns warning ──
    if unknown_columns:
        warnings.append(IngestWarning(
            code="UNKNOWN_COLUMNS",
            message=(
                f"Found {len(unknown_columns)} unknown column(s) not in the standard "
                f"Prowler schema. Stored in extra_fields: {unknown_columns}. "
                "This may indicate a newer Prowler version."
            ),
        ))

    # ── Determine scanner version from findings ──
    scanner_version = ""
    for f in findings:
        if f.raw_prowler_version:
            scanner_version = f.raw_prowler_version
            break

    logger.info(
        "Ingestion complete: %d findings parsed, %d warnings",
        len(findings), len(warnings),
    )

    return IngestResult(
        run_id=run_id,
        source_file=str(path),
        source_file_hash=file_hash,
        scanner="prowler",
        scanner_version=scanner_version,
        sheet_name=sheet_name,
        total_rows_read=len(row_dicts),
        findings=findings,
        warnings=warnings,
        unknown_columns=unknown_columns,
    )