"""
stage5_render_excel.py — Stage 5: Excel Report Renderer

Renders the final client-facing Excel report from enriched groups.

Contract:
    render_excel(enrich_result, config, output_dir) -> Path

Column mapping (matches Output_Template.xlsx exactly):
    A  Ref              — AWS1, AWS2... assigned at render time
    B  Finding          — finding_title from LLM
    C  Risk Rating      — colour-coded High/Medium/Low/Critical
    D  Root Cause       — root_cause_narrative from LLM
    E  Likelihood       — likelihood_rating from Stage 2
    F  Consequence      — consequence_rating from LLM
    G  Access Required  — access_required from LLM
    H  Situation        — situation_narrative from LLM
    I  Consequence      — consequence_narrative from LLM
    J  Recommendations  — remediation_recommendation_text from raw finding

Each finding row also appends a sub-row listing all affected resources
(ARNs / names) so the report is fully traceable.

Formula injection defence:
    Every string value is sanitised — cells starting with = + - @ are
    prefixed with a single quote to prevent formula execution.
"""

from __future__ import annotations

import re
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Optional

import openpyxl
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))

from models import CanonicalFinding
from stage2_5_grouping import GroupedOutputGroup
from stage3_llm import EnrichResult

# ── Constants ─────────────────────────────────────────────────────────

FONT_NAME  = "Aptos Narrow"
FONT_SIZE  = 11

# Column indices (1-based)
COL_REF         = 1
COL_FINDING     = 2
COL_RISK        = 3
COL_ROOT_CAUSE  = 4
COL_LIKELIHOOD  = 5
COL_CONSEQUENCE = 6
COL_ACCESS      = 7
COL_SITUATION   = 8
COL_CONSEQUENCE_NARRATIVE = 9
COL_RECOMMENDATIONS = 10

TOTAL_COLS = 10

# Risk rating colours (ARGB)
RISK_COLOURS = {
    "Critical": "FF7030A0",  # purple
    "High":     "FFFF0000",  # red
    "Medium":   "FFFFC000",  # amber
    "Low":      "FF92D050",  # green
}

# Approximate column widths (characters)
COL_WIDTHS = {
    COL_REF:                  8,
    COL_FINDING:              35,
    COL_RISK:                 14,
    COL_ROOT_CAUSE:           45,
    COL_LIKELIHOOD:           16,
    COL_CONSEQUENCE:          18,
    COL_ACCESS:               35,
    COL_SITUATION:            50,
    COL_CONSEQUENCE_NARRATIVE:45,
    COL_RECOMMENDATIONS:      50,
}

# Formula injection prefixes to sanitise
_FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r")


# ── Sanitisation ──────────────────────────────────────────────────────

def _safe(value: Any) -> str:
    """
    Convert to string and strip formula injection characters.
    Cells starting with = + - @ are prefixed with a space — Excel
    then treats them as text, not formulas.
    """
    if value is None:
        return ""
    s = str(value).strip()
    if s and s[0] in _FORMULA_PREFIXES:
        s = " " + s
    return s


# ── Style helpers ─────────────────────────────────────────────────────

def _font(bold: bool = False, size: int = FONT_SIZE, color: str = "FF000000") -> Font:
    return Font(name=FONT_NAME, size=size, bold=bold, color=color)


def _fill(rgb: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=rgb)


def _align(
    horizontal: str = "left",
    vertical:   str = "top",
    wrap:       bool = True,
) -> Alignment:
    return Alignment(
        horizontal=horizontal,
        vertical=vertical,
        wrap_text=wrap,
    )


def _thin_border() -> Border:
    thin = Side(style="thin", color="FFD0D0D0")
    return Border(top=thin, bottom=thin, left=thin, right=thin)


# ── Cell writer ───────────────────────────────────────────────────────

def _write(
    ws,
    row:   int,
    col:   int,
    value: Any,
    bold:  bool = False,
    fill:  Optional[str] = None,
    halign: str = "left",
    valign: str = "top",
    wrap:  bool = True,
    font_color: str = "FF000000",
    font_size:  int = FONT_SIZE,
) -> None:
    cell = ws.cell(row=row, column=col, value=_safe(value))
    cell.font      = _font(bold=bold, size=font_size, color=font_color)
    cell.alignment = _align(horizontal=halign, vertical=valign, wrap=wrap)
    cell.border    = _thin_border()
    if fill:
        cell.fill = _fill(fill)


# ── Header row ────────────────────────────────────────────────────────

def _write_header(ws, row: int = 1) -> None:
    headers = [
        (COL_REF,               "Ref",               "center"),
        (COL_FINDING,           "Finding",           "left"),
        (COL_RISK,              "Risk Rating",       "center"),
        (COL_ROOT_CAUSE,        "Root Cause",        "left"),
        (COL_LIKELIHOOD,        "Likelihood Rating", "left"),
        (COL_CONSEQUENCE,       "Consequence Rating","left"),
        (COL_ACCESS,            "Access Required",   "left"),
        (COL_SITUATION,         "Situation",         "left"),
        (COL_CONSEQUENCE_NARRATIVE, "Consequence",   "left"),
        (COL_RECOMMENDATIONS,   "Recommendations",   "left"),
    ]
    for col, label, halign in headers:
        cell = ws.cell(row=row, column=col, value=label)
        cell.font      = _font(bold=True)
        cell.alignment = _align(horizontal=halign, vertical="center", wrap=True)
        cell.fill      = _fill("FFD9D9D9")  # light grey header
        cell.border    = _thin_border()

    ws.row_dimensions[row].height = 43.2


# ── Section header ────────────────────────────────────────────────────

def _write_section_header(ws, row: int, label: str) -> None:
    """Write a full-width bold section header row (e.g. 'AWS')."""
    cell = ws.cell(row=row, column=1, value=label)
    cell.font      = _font(bold=True)
    cell.alignment = _align(horizontal="left", vertical="center", wrap=True)
    cell.fill      = _fill("FFE0E0E0")
    cell.border    = _thin_border()

    # Merge across all 10 columns
    ws.merge_cells(
        start_row=row, start_column=1,
        end_row=row,   end_column=TOTAL_COLS,
    )
    ws.row_dimensions[row].height = 18


# ── Resource detail row ───────────────────────────────────────────────

def _build_resource_text(
    all_findings: list[CanonicalFinding],
    instance_ids: list[str],
) -> str:
    """
    Build the affected-resources text block for embedding inside the
    Situation column (not a separate row).
    """
    findings_map = {f.finding_instance_id: f for f in all_findings}
    resources: list[str] = []

    for fid in instance_ids:
        f = findings_map.get(fid)
        if not f:
            continue
        res = (
            f.resource_uid_normalised
            or f.raw_resource_name
            or f.raw_resource_uid
            or ""
        )
        acct = f.raw_account_name or f.raw_account_uid or ""
        if res and res not in ("", "no_resource"):
            entry = res
            if acct:
                entry = entry + " (" + acct + ")"
            if entry not in resources:
                resources.append(entry)

    if not resources:
        return ""

    lines = ["Affected resources:"]
    for r in resources:
        lines.append("  \u2022 " + r)
    return "\n".join(lines)


# ── Finding row ───────────────────────────────────────────────────────

def _build_recommendations(group) -> str:
    source_groups = getattr(group, "source_groups", None)
    if source_groups and len(source_groups) > 1:
        parts = []
        for sg in source_groups:
            rep   = sg.representative
            title = rep.raw_check_title or sg.check_id
            remed = (rep.raw_remediation_recommendation_text or "").strip()
            cli   = (rep.raw_remediation_code_cli or "").strip()
            tf    = (rep.raw_remediation_code_terraform or "").strip()
            body  = remed if remed else "[See vendor documentation]"
            if cli:
                body = body + "\n  CLI: " + cli
            if tf:
                body = body + "\n  Terraform: " + tf
            parts.append(title + ":\n" + body)
        return "\n\n".join(parts)
    else:
        rep   = group.representative
        remed = (rep.raw_remediation_recommendation_text or "").strip()
        cli   = (rep.raw_remediation_code_cli or "").strip()
        tf    = (rep.raw_remediation_code_terraform or "").strip()
        lines = []
        if remed:
            lines.append(remed)
        if cli:
            lines.append("CLI: " + cli)
        if tf:
            lines.append("Terraform: " + tf)
        return "\n".join(lines)


def _write_finding_row(
    ws,
    row:         int,
    ref:         str,
    group:       GroupedOutputGroup,
    all_findings: list[CanonicalFinding],
) -> None:
    """Write one finding row from a GroupedOutputGroup."""
    rep             = group.representative
    risk_rating     = rep.risk_rating or "Medium"
    risk_colour     = RISK_COLOURS.get(risk_rating, "FFFFFFFF")
    recommendations = _build_recommendations(group)

    # Embed affected resources into the Situation narrative
    resource_text = _build_resource_text(all_findings, group.instance_ids)
    situation = rep.situation_narrative or ""
    if resource_text:
        situation = situation.rstrip() + "\n\n" + resource_text

    _write(ws, row, COL_REF,          ref,                       bold=False, halign="center", valign="center")
    _write(ws, row, COL_FINDING,      rep.finding_title          or group.group_name)
    _write(ws, row, COL_RISK,         risk_rating,               bold=True,  halign="center", valign="center", fill=risk_colour, font_color="FFFFFFFF")
    _write(ws, row, COL_ROOT_CAUSE,   rep.root_cause_narrative   or "")
    _write(ws, row, COL_LIKELIHOOD,   group.likelihood_rating    or "")
    _write(ws, row, COL_CONSEQUENCE,  rep.consequence_rating     or "")
    _write(ws, row, COL_ACCESS,       rep.access_required        or "")
    _write(ws, row, COL_SITUATION,    situation)
    _write(ws, row, COL_CONSEQUENCE_NARRATIVE, rep.consequence_narrative or "")
    _write(ws, row, COL_RECOMMENDATIONS, recommendations)

    remed_lines = recommendations.count("\n") + 1
    sit_lines   = situation.count("\n") + 1
    ws.row_dimensions[row].height = max(80, min(max(remed_lines, sit_lines) * 15, 220))


# ── Column widths ─────────────────────────────────────────────────────

def _set_column_widths(ws) -> None:
    for col, width in COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col)].width = width


# ── Freeze panes ──────────────────────────────────────────────────────

def _freeze_header(ws) -> None:
    """Freeze the header row so it stays visible when scrolling."""
    ws.freeze_panes = "A2"


# ── Main renderer ─────────────────────────────────────────────────────

def render_excel(
    enrich_result: EnrichResult,
    config:        dict[str, Any],
    output_dir:    Path,
    template_path: Optional[Path] = None,
) -> Path:
    """
    Render the final Excel report.

    Args:
        enrich_result:  Output of Stage 3 enrich_grouped().
        config:         Loaded config dict.
        output_dir:     Directory to write the report into.
        template_path:  Optional path to Output_Template.xlsx.
                        If not provided, looks in config then falls back
                        to a clean workbook.

    Returns:
        Path to the written .xlsx file.
    """
    engagement  = config.get("engagement", {})
    client_name = engagement.get("client_name", "Client")
    period      = engagement.get("assessment_period", "")
    ref_prefix  = config.get("output", {}).get("ref_prefix", "AWS")
    filename    = engagement.get("output_filename") or f"SecurityReport_{period.replace(' ','')}.xlsx"

    output_path = output_dir / filename

    # ── Load template or create fresh workbook ──
    tmpl = template_path
    if tmpl is None:
        cfg_tmpl = config.get("output", {}).get("template_path", "")
        if cfg_tmpl:
            tmpl = Path(cfg_tmpl)
            if not tmpl.is_absolute():
                # Resolve relative to project root (two levels up from src/)
                tmpl = Path(__file__).resolve().parent.parent / tmpl

    if tmpl and tmpl.exists():
        wb = openpyxl.load_workbook(str(tmpl))
        # Use AWS sheet from template; clear data rows but keep header
        if "AWS" in wb.sheetnames:
            ws = wb["AWS"]
            # Delete all rows below header
            if ws.max_row > 1:
                ws.delete_rows(2, ws.max_row - 1)
            # Remove any merged cells below row 1
            for mr in list(ws.merged_cells.ranges):
                if mr.min_row > 1:
                    ws.merged_cells.remove(mr)
        else:
            ws = wb.create_sheet("AWS")
            _write_header(ws)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "AWS"
        _write_header(ws)

    _set_column_widths(ws)
    _freeze_header(ws)

    # ── Write findings ──
    current_row = 2
    ref_counter = 1

    # Only groups for the AWS section
    aws_groups = [
        g for g in enrich_result.output_groups
        if g.output_section == "AWS"
    ]

    if not aws_groups:
        # Write a placeholder if no groups
        ws.cell(row=current_row, column=1, value="No findings in this section.")
    else:
        for group in aws_groups:
            ref = f"{ref_prefix}{ref_counter}"
            ref_counter += 1

            # Finding row — resources are embedded in the Situation column
            _write_finding_row(ws, current_row, ref, group, enrich_result.all_findings)
            current_row += 1

    # ── Run info in a note at the top of a metadata sheet ──
    if "_meta" not in wb.sheetnames:
        meta_ws = wb.create_sheet("_meta")
    else:
        meta_ws = wb["_meta"]

    meta_ws["A1"] = "Run ID"
    meta_ws["B1"] = enrich_result.run_id
    meta_ws["A2"] = "Generated"
    meta_ws["B2"] = datetime.now(timezone.utc).isoformat()
    meta_ws["A3"] = "Client"
    meta_ws["B3"] = client_name
    meta_ws["A4"] = "Period"
    meta_ws["B4"] = period
    meta_ws["A5"] = "Groups"
    meta_ws["B5"] = enrich_result.group_count
    meta_ws["A6"] = "Enriched"
    meta_ws["B6"] = enrich_result.enriched_count
    meta_ws["A7"] = "LLM Failed"
    meta_ws["B7"] = enrich_result.failed_count
    meta_ws.sheet_state = "hidden"

    wb.save(str(output_path))

    print(
        f"  Excel report written: {output_path.name} "
        f"({output_path.stat().st_size // 1024} KB)",
        flush=True,
    )
    return output_path