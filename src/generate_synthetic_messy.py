"""
generate_synthetic_messy.py — Realistic messy synthetic Prowler data.

Unlike the clean synthetic, this file mimics what real Prowler output actually
looks like: inconsistent blanks, partial fields, mixed account states, and the
exact blank patterns seen in production scans.

Blank patterns reproduced:
    - DESCRIPTION blank on some checks (community checks, older check versions)
    - RISK blank on some checks (same cause)
    - REMEDIATION_RECOMMENDATION_TEXT blank on custom checks
    - REMEDIATION_CODE_* almost entirely blank (normal)
    - SUBSERVICE_NAME blank on most rows (normal)
    - RESOURCE_DETAILS blank on most rows (normal)
    - ACCOUNT_ORGANIZATION_UID / NAME blank (account not in AWS Org)
    - ACCOUNT_TAGS blank (most accounts have none)
    - RESOURCE_TAGS blank (most resources untagged)
    - REGION blank for IAM/account-level checks
    - RELATED_URL blank on some checks
    - NOTES always blank (analyst field)
    - DEPENDS_ON / RELATED_TO almost always blank
    - STATUS_EXTENDED sometimes terse, sometimes detailed, sometimes blank

Accounts:
    - Three accounts: prod, dev, legacy
    - Legacy account NOT in an AWS Org (ACCOUNT_ORGANIZATION_* blank)
    - Only prod account has ACCOUNT_TAGS

Checks covered (35 rows, 14 distinct check_ids):
    - s3_bucket_public_access (5 resources, 2 accounts, 1 duplicate)
    - s3_bucket_default_encryption (2 resources)
    - s3_bucket_versioning_enabled (PASS — excluded downstream)
    - iam_user_mfa_enabled_console_access (3 users, blank DESCRIPTION)
    - iam_root_mfa_enabled (account singleton, no resource)
    - iam_user_access_key_age_90_days (MUTED=True, STATUS=FAIL mismatch)
    - iam_password_policy_lowercase (blank RISK + blank REMEDIATION)
    - ec2_securitygroup_allow_ingress_ssh (critical, internet-exposed)
    - ec2_instance_imdsv2_enabled (name-only, no ARN)
    - ec2_ebs_volume_encryption (3 volumes, blank STATUS_EXTENDED on one)
    - rds_instance_storage_encrypted (blank DESCRIPTION + RISK)
    - rds_instance_no_public_access (exact duplicate pair — dedup test)
    - vpc_flow_logs_enabled (medium, internet-exposed override test)
    - cloudtrail_multi_region_enabled (blank REMEDIATION_CODE_*)
    - guardduty_is_enabled (MANUAL status)
    - securityhub_enabled (future column present)
    - config_recorder_all_regions_enabled (account-level, no resource)
    - cloudwatch_log_group_retention_policy (informational severity)
"""

from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path

import openpyxl
from openpyxl.styles import Font

# ── Constants ─────────────────────────────────────────────────────────

HEADERS = [
    "AUTH_METHOD", "TIMESTAMP", "ACCOUNT_UID", "ACCOUNT_NAME", "ACCOUNT_EMAIL",
    "ACCOUNT_ORGANIZATION_UID", "ACCOUNT_ORGANIZATION_NAME", "ACCOUNT_TAGS",
    "FINDING_UID", "PROVIDER", "CHECK_ID", "CHECK_TITLE", "CHECK_TYPE",
    "STATUS", "STATUS_EXTENDED", "MUTED", "SERVICE_NAME", "SUBSERVICE_NAME",
    "SEVERITY", "RESOURCE_TYPE", "RESOURCE_UID", "RESOURCE_NAME",
    "RESOURCE_DETAILS", "RESOURCE_TAGS", "PARTITION", "REGION",
    "DESCRIPTION", "RISK", "RELATED_URL", "REMEDIATION_RECOMMENDATION_TEXT",
    "REMEDIATION_RECOMMENDATION_URL", "REMEDIATION_CODE_NATIVEIAC",
    "REMEDIATION_CODE_TERRAFORM", "REMEDIATION_CODE_CLI", "REMEDIATION_CODE_OTHER",
    "COMPLIANCE", "CATEGORIES", "DEPENDS_ON", "RELATED_TO", "NOTES",
    "PROWLER_VERSION",
]

TS = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
PROWLER_VER = "4.3.1"
PARTITION = "aws"
PROVIDER = "aws"

# ── Three accounts ────────────────────────────────────────────────────

PROD = {
    "uid":   "111122223333",
    "name":  "acme-prod",
    "email": "aws-prod@acme.example",
    "org_uid":  "o-acmeorg001",
    "org_name": "acme-org",
    "tags":  "Environment:prod | CostCentre:platform | Terraform:true",
}

DEV = {
    "uid":   "444455556666",
    "name":  "acme-dev",
    "email": "aws-dev@acme.example",
    "org_uid":  "o-acmeorg001",
    "org_name": "acme-org",
    "tags":  None,   # ← dev account has no ACCOUNT_TAGS
}

LEGACY = {
    "uid":   "777788889999",
    "name":  "acme-legacy",
    "email": "aws-legacy@acme.example",
    "org_uid":  None,  # ← NOT in AWS Org — structural blank
    "org_name": None,  # ← structural blank
    "tags":  None,
}


def make_finding_uid(check_id: str, account_uid: str, region: str, resource: str) -> str:
    region_part = region or "global"
    res_part = (resource or "nores")[:20]
    return f"prowler-aws-{check_id[:20]}-{account_uid}-{region_part}-{res_part}"


def row(
    acct: dict,
    check_id: str,
    check_title: str,
    check_type: str,
    status: str,
    status_extended: str | None,
    muted: str,
    service_name: str,
    severity: str,
    resource_type: str,
    resource_uid: str | None,
    resource_name: str | None,
    region: str | None,
    *,
    subservice_name: str | None = None,
    resource_details: str | None = None,
    resource_tags: str | None = None,
    description: str | None = None,
    risk: str | None = None,
    related_url: str | None = None,
    remediation_text: str | None = None,
    remediation_url: str | None = None,
    remediation_nativeiac: str | None = None,
    remediation_terraform: str | None = None,
    remediation_cli: str | None = None,
    remediation_other: str | None = None,
    compliance: str | None = None,
    categories: str | None = None,
    depends_on: str | None = None,
    related_to: str | None = None,
    notes: str | None = None,
) -> dict:
    resource_key = resource_uid or resource_name or "nores"
    return {
        "AUTH_METHOD":                      "profile:None",
        "TIMESTAMP":                        TS,
        "ACCOUNT_UID":                      acct["uid"],
        "ACCOUNT_NAME":                     acct["name"],
        "ACCOUNT_EMAIL":                    acct["email"],
        "ACCOUNT_ORGANIZATION_UID":         acct["org_uid"],   # None for LEGACY
        "ACCOUNT_ORGANIZATION_NAME":        acct["org_name"],  # None for LEGACY
        "ACCOUNT_TAGS":                     acct["tags"],      # None for DEV + LEGACY
        "FINDING_UID":                      make_finding_uid(check_id, acct["uid"], region or "global", resource_key),
        "PROVIDER":                         PROVIDER,
        "CHECK_ID":                         check_id,
        "CHECK_TITLE":                      check_title,
        "CHECK_TYPE":                       check_type,
        "STATUS":                           status,
        "STATUS_EXTENDED":                  status_extended,   # sometimes blank
        "MUTED":                            muted,
        "SERVICE_NAME":                     service_name,
        "SUBSERVICE_NAME":                  subservice_name,   # almost always None
        "SEVERITY":                         severity,
        "RESOURCE_TYPE":                    resource_type,
        "RESOURCE_UID":                     resource_uid,
        "RESOURCE_NAME":                    resource_name,
        "RESOURCE_DETAILS":                 resource_details,  # almost always None
        "RESOURCE_TAGS":                    resource_tags,     # often None
        "PARTITION":                        PARTITION,
        "REGION":                           region,
        "DESCRIPTION":                      description,       # sometimes blank
        "RISK":                             risk,              # sometimes blank
        "RELATED_URL":                      related_url,
        "REMEDIATION_RECOMMENDATION_TEXT":  remediation_text,  # sometimes blank
        "REMEDIATION_RECOMMENDATION_URL":   remediation_url,
        "REMEDIATION_CODE_NATIVEIAC":       remediation_nativeiac,   # almost always None
        "REMEDIATION_CODE_TERRAFORM":       remediation_terraform,   # almost always None
        "REMEDIATION_CODE_CLI":             remediation_cli,
        "REMEDIATION_CODE_OTHER":           remediation_other,       # almost always None
        "COMPLIANCE":                       compliance,
        "CATEGORIES":                       categories,
        "DEPENDS_ON":                       depends_on,        # almost always None
        "RELATED_TO":                       related_to,        # almost always None
        "NOTES":                            notes,             # always None
        "PROWLER_VERSION":                  PROWLER_VER,
    }


def build_rows() -> list[dict]:
    rows: list[dict] = []

    # ═══════════════════════════════════════════════════════════
    # S3 checks
    # ═══════════════════════════════════════════════════════════

    # 1. S3 public access — PROD — bucket with ARN — full fields
    rows.append(row(
        PROD,
        check_id="s3_bucket_public_access",
        check_title="Ensure S3 bucket does not allow public access",
        check_type="Software and Configuration Checks",
        status="FAIL",
        status_extended="Bucket acme-prod-logs has Block Public Access disabled.",
        muted="False",
        service_name="s3",
        severity="high",
        resource_type="AWSS3Bucket",
        resource_uid="arn:aws:s3:::acme-prod-logs",
        resource_name="acme-prod-logs",
        region="ap-southeast-2",
        resource_tags="Project:logging | Owner:platform",
        description="This control checks whether Amazon S3 buckets have Block Public Access enabled.",
        risk="Public S3 buckets expose data to the internet. Attackers can exfiltrate sensitive data.",
        related_url="https://docs.aws.amazon.com/AmazonS3/latest/userguide/access-control-block-public-access.html",
        remediation_text="Enable S3 Block Public Access settings at the bucket level for all four settings.",
        remediation_url="https://docs.aws.amazon.com/AmazonS3/latest/userguide/access-control-block-public-access.html",
        remediation_cli="aws s3api put-public-access-block --bucket acme-prod-logs "
                        "--public-access-block-configuration BlockPublicAcls=true,IgnorePublicAcls=true,"
                        "BlockPublicPolicy=true,RestrictPublicBuckets=true",
        compliance="CIS-2.0 | 2.1.5 | NIST-800-53 | SC-7 | PCI-DSS | 1.2",
        categories="internet-exposed | data-protection",
    ))

    # 2. S3 public access — PROD — second bucket — status_extended terse
    rows.append(row(
        PROD,
        check_id="s3_bucket_public_access",
        check_title="Ensure S3 bucket does not allow public access",
        check_type="Software and Configuration Checks",
        status="FAIL",
        status_extended="Bucket acme-prod-backups is public.",  # terse
        muted="False",
        service_name="s3",
        severity="high",
        resource_type="AWSS3Bucket",
        resource_uid="arn:aws:s3:::acme-prod-backups",
        resource_name="acme-prod-backups",
        region="ap-southeast-2",
        resource_tags=None,  # ← untagged bucket
        description="This control checks whether Amazon S3 buckets have Block Public Access enabled.",
        risk="Public S3 buckets expose data to the internet. Attackers can exfiltrate sensitive data.",
        related_url="https://docs.aws.amazon.com/AmazonS3/latest/userguide/access-control-block-public-access.html",
        remediation_text="Enable S3 Block Public Access settings at the bucket level for all four settings.",
        remediation_url="https://docs.aws.amazon.com/AmazonS3/latest/userguide/access-control-block-public-access.html",
        remediation_cli="aws s3api put-public-access-block --bucket acme-prod-backups "
                        "--public-access-block-configuration BlockPublicAcls=true,IgnorePublicAcls=true,"
                        "BlockPublicPolicy=true,RestrictPublicBuckets=true",
        compliance="CIS-2.0 | 2.1.5",
        categories="internet-exposed | data-protection",
    ))

    # 3. S3 public access — DEV account — no ACCOUNT_TAGS
    rows.append(row(
        DEV,
        check_id="s3_bucket_public_access",
        check_title="Ensure S3 bucket does not allow public access",
        check_type="Software and Configuration Checks",
        status="FAIL",
        status_extended=None,   # ← blank STATUS_EXTENDED (real Prowler behaviour)
        muted="False",
        service_name="s3",
        severity="high",
        resource_type="AWSS3Bucket",
        resource_uid="arn:aws:s3:::acme-dev-uploads",
        resource_name="acme-dev-uploads",
        region="ap-southeast-2",
        description="This control checks whether Amazon S3 buckets have Block Public Access enabled.",
        risk="Public S3 buckets expose data to the internet.",
        remediation_text="Enable S3 Block Public Access settings.",
        remediation_cli="aws s3api put-public-access-block --bucket acme-dev-uploads "
                        "--public-access-block-configuration BlockPublicAcls=true,IgnorePublicAcls=true,"
                        "BlockPublicPolicy=true,RestrictPublicBuckets=true",
        compliance="CIS-2.0 | 2.1.5",
        categories="internet-exposed",
    ))

    # 4. S3 public access — LEGACY account — NOT in AWS Org
    rows.append(row(
        LEGACY,
        check_id="s3_bucket_public_access",
        check_title="Ensure S3 bucket does not allow public access",
        check_type="Software and Configuration Checks",
        status="FAIL",
        status_extended="Bucket legacy-data is publicly accessible.",
        muted="False",
        service_name="s3",
        severity="high",
        resource_type="AWSS3Bucket",
        resource_uid="arn:aws:s3:::legacy-data",
        resource_name="legacy-data",
        region="us-east-1",
        description="This control checks whether Amazon S3 buckets have Block Public Access enabled.",
        risk="Public S3 buckets expose data to the internet.",
        remediation_text="Enable S3 Block Public Access settings.",
        compliance="CIS-2.0 | 2.1.5",
        categories="internet-exposed",
    ))

    # 5. S3 public access — PROD — exact duplicate of row 1 (dedup test)
    rows.append(row(
        PROD,
        check_id="s3_bucket_public_access",
        check_title="Ensure S3 bucket does not allow public access",
        check_type="Software and Configuration Checks",
        status="FAIL",
        status_extended="Bucket acme-prod-logs has Block Public Access disabled.",
        muted="False",
        service_name="s3",
        severity="high",
        resource_type="AWSS3Bucket",
        resource_uid="arn:aws:s3:::acme-prod-logs",   # ← same ARN as row 1
        resource_name="acme-prod-logs",
        region="ap-southeast-2",
        resource_tags="Project:logging | Owner:platform",
        description="This control checks whether Amazon S3 buckets have Block Public Access enabled.",
        risk="Public S3 buckets expose data to the internet.",
        remediation_text="Enable S3 Block Public Access settings.",
        compliance="CIS-2.0 | 2.1.5",
        categories="internet-exposed | data-protection",
    ))

    # 6. S3 encryption — MUTED(FAIL) — bucket has been accepted risk
    rows.append(row(
        PROD,
        check_id="s3_bucket_default_encryption",
        check_title="Ensure S3 bucket has default encryption enabled",
        check_type="Software and Configuration Checks",
        status="MUTED(FAIL)",
        status_extended="Bucket acme-prod-archive has no default encryption configured.",
        muted="True",
        service_name="s3",
        severity="medium",
        resource_type="AWSS3Bucket",
        resource_uid="arn:aws:s3:::acme-prod-archive",
        resource_name="acme-prod-archive",
        region="ap-southeast-2",
        description="This control checks whether default encryption is enabled on Amazon S3 buckets.",
        risk="Without default encryption, objects are stored in plaintext.",
        remediation_text="Enable default encryption on the S3 bucket using SSE-S3 or SSE-KMS.",
        compliance="CIS-2.0 | 2.1.1",
        categories="encryption",
    ))

    # 7. S3 versioning — PASS (excluded downstream)
    rows.append(row(
        PROD,
        check_id="s3_bucket_versioning_enabled",
        check_title="Ensure S3 bucket has versioning enabled",
        check_type="Software and Configuration Checks",
        status="PASS",
        status_extended="Bucket acme-prod-logs has versioning enabled.",
        muted="False",
        service_name="s3",
        severity="low",
        resource_type="AWSS3Bucket",
        resource_uid="arn:aws:s3:::acme-prod-logs",
        resource_name="acme-prod-logs",
        region="ap-southeast-2",
        description="Checks whether S3 bucket versioning is enabled.",
        risk=None,  # PASS findings sometimes have no risk
        remediation_text=None,  # PASS — no remediation needed
        categories="resilience",
    ))

    # ═══════════════════════════════════════════════════════════
    # IAM checks  (global — REGION intentionally blank)
    # ═══════════════════════════════════════════════════════════

    # 8. IAM MFA — 3 users — DESCRIPTION blank on all (community check)
    for username in ["john.smith", "jane.doe", "svc-deploy"]:
        rows.append(row(
            PROD,
            check_id="iam_user_mfa_enabled_console_access",
            check_title="Ensure MFA is enabled for all IAM users with console access",
            check_type="Software and Configuration Checks",
            status="FAIL",
            status_extended=f"IAM user {username} has console access but no MFA device.",
            muted="False",
            service_name="iam",
            severity="critical",
            resource_type="AWSIAMUser",
            resource_uid=f"arn:aws:iam::111122223333:user/{username}",
            resource_name=username,
            region=None,      # ← global IAM check, structural blank
            description=None, # ← blank DESCRIPTION (Category 2 data quality)
            risk=None,        # ← blank RISK (Category 2 data quality)
            related_url="https://docs.aws.amazon.com/IAM/latest/UserGuide/id_credentials_mfa.html",
            remediation_text="Enable MFA for the IAM user via the AWS Console or CLI.",
            remediation_cli=f"aws iam enable-mfa-device --user-name {username} --serial-number <mfa-arn> --authentication-code1 <code1> --authentication-code2 <code2>",
            compliance="CIS-2.0 | 1.10 | NIST-800-53 | IA-2",
            categories="identity-management | mfa",
        ))

    # 11. IAM root MFA — account singleton, no resource at all
    r = row(
        PROD,
        check_id="iam_root_mfa_enabled",
        check_title="Ensure MFA is enabled for the root account",
        check_type="Software and Configuration Checks",
        status="FAIL",
        status_extended="The root account does not have an MFA device configured.",
        muted="False",
        service_name="iam",
        severity="critical",
        resource_type="AWSIAMAccountSummary",
        resource_uid=None,  # ← no ARN for account-level
        resource_name=None, # ← no resource name
        region=None,
        description="Checks whether MFA is enabled for the AWS account root user.",
        risk="The root account has unrestricted access to all AWS services. Without MFA it is vulnerable to credential theft.",
        related_url="https://docs.aws.amazon.com/IAM/latest/UserGuide/id_root-user.html",
        remediation_text="Enable a hardware or virtual MFA device for the root account.",
        compliance="CIS-2.0 | 1.5 | NIST-800-53 | IA-2(1)",
        categories="identity-management | mfa",
    )
    r["RESOURCE_UID"] = None
    r["RESOURCE_NAME"] = None
    rows.append(r)

    # 12. IAM access key rotation — MUTED=True, STATUS=FAIL (reconciliation test)
    rows.append(row(
        PROD,
        check_id="iam_user_access_key_age_90_days",
        check_title="Ensure IAM user access keys are rotated every 90 days or less",
        check_type="Software and Configuration Checks",
        status="FAIL",       # ← STATUS says FAIL
        status_extended="Access key AKIAI... for user svc-deploy is 134 days old.",
        muted="True",        # ← MUTED=True — reconcile to MUTED(FAIL)
        service_name="iam",
        severity="medium",
        resource_type="AWSIAMUser",
        resource_uid="arn:aws:iam::111122223333:user/svc-deploy",
        resource_name="svc-deploy",
        region=None,
        description="This control checks whether IAM user access keys are rotated within 90 days.",
        risk="Long-lived access keys increase the risk of credential compromise.",
        remediation_text="Rotate IAM user access keys and delete old keys.",
        compliance="CIS-2.0 | 1.14",
        categories="identity-management | credential-hygiene",
    ))

    # 13. IAM password policy — blank RISK and blank REMEDIATION_TEXT
    rows.append(row(
        PROD,
        check_id="iam_password_policy_lowercase",
        check_title="Ensure IAM password policy requires at least one lowercase letter",
        check_type="Software and Configuration Checks",
        status="FAIL",
        status_extended="Password policy does not require lowercase letters.",
        muted="False",
        service_name="iam",
        severity="medium",
        resource_type="AWSIAMPasswordPolicy",
        resource_uid="arn:aws:iam::111122223333:account",
        resource_name="password-policy",
        region=None,
        description="Checks whether the IAM password policy requires lowercase letters.",
        risk=None,                  # ← blank RISK (Category 2)
        related_url=None,           # ← blank RELATED_URL (by design)
        remediation_text=None,      # ← blank REMEDIATION_TEXT (Category 2)
        remediation_url=None,
        compliance="CIS-2.0 | 1.8",
        categories="identity-management",
    ))

    # ═══════════════════════════════════════════════════════════
    # EC2 checks
    # ═══════════════════════════════════════════════════════════

    # 14. EC2 SSH open — critical, internet-exposed
    rows.append(row(
        PROD,
        check_id="ec2_securitygroup_allow_ingress_from_internet_to_tcp_port_22",
        check_title="Ensure no security groups allow ingress from 0.0.0.0/0 to port 22",
        check_type="Software and Configuration Checks",
        status="FAIL",
        status_extended="Security group sg-0prod1234 (web-servers) allows unrestricted SSH (0.0.0.0/0:22).",
        muted="False",
        service_name="ec2",
        severity="critical",
        resource_type="AWSEC2SecurityGroup",
        resource_uid="arn:aws:ec2:ap-southeast-2:111122223333:security-group/sg-0prod1234",
        resource_name="sg-0prod1234",
        region="ap-southeast-2",
        resource_tags="Name:web-servers | Environment:prod",
        description="Checks whether security groups allow unrestricted inbound access on TCP port 22 (SSH).",
        risk="Unrestricted SSH access allows anyone to attempt brute-force login to EC2 instances.",
        related_url="https://docs.aws.amazon.com/vpc/latest/userguide/VPC_SecurityGroups.html",
        remediation_text="Restrict SSH ingress to specific trusted IP ranges. Remove the 0.0.0.0/0 rule.",
        remediation_cli="aws ec2 revoke-security-group-ingress --group-id sg-0prod1234 --protocol tcp --port 22 --cidr 0.0.0.0/0",
        compliance="CIS-2.0 | 5.2 | NIST-800-53 | SC-7 | PCI-DSS | 1.2",
        categories="internet-exposed | network-security",
    ))

    # 15. EC2 IMDSv2 — name only, no ARN (ARN fallback test)
    rows.append(row(
        PROD,
        check_id="ec2_instance_imdsv2_enabled",
        check_title="Ensure EC2 instances use IMDSv2",
        check_type="Software and Configuration Checks",
        status="FAIL",
        status_extended="Instance i-0aabbccdd does not enforce IMDSv2.",
        muted="False",
        service_name="ec2",
        severity="high",
        resource_type="AWSEC2Instance",
        resource_uid=None,          # ← no ARN — ARN fallback test
        resource_name="i-0aabbccdd",
        region="ap-southeast-2",
        description="Checks whether EC2 instances are configured to use Instance Metadata Service Version 2.",
        risk="IMDSv1 is susceptible to SSRF attacks that can expose IAM credentials.",
        remediation_text="Enforce IMDSv2 on all EC2 instances by setting HttpTokens to required.",
        remediation_cli="aws ec2 modify-instance-metadata-options --instance-id i-0aabbccdd --http-tokens required",
        compliance="CIS-2.0 | 5.6",
        categories="configuration | credential-protection",
    ))

    # 16–18. EBS encryption — 3 volumes, one with blank STATUS_EXTENDED
    ebs_vols = [
        ("vol-0aaa1111", "ap-southeast-2a", "EBS volume vol-0aaa1111 is not encrypted."),
        ("vol-0bbb2222", "ap-southeast-2b", None),   # ← blank STATUS_EXTENDED
        ("vol-0ccc3333", "ap-southeast-2a", "EBS volume vol-0ccc3333 is not encrypted."),
    ]
    for vol_id, az, se in ebs_vols:
        rows.append(row(
            PROD,
            check_id="ec2_ebs_volume_encryption_enabled",
            check_title="Ensure EBS volumes are encrypted",
            check_type="Software and Configuration Checks",
            status="FAIL",
            status_extended=se,
            muted="False",
            service_name="ec2",
            severity="high",
            resource_type="AWSEC2Volume",
            resource_uid=f"arn:aws:ec2:ap-southeast-2:111122223333:volume/{vol_id}",
            resource_name=vol_id,
            region="ap-southeast-2",
            description="Checks whether EBS volumes are encrypted.",
            risk="Unencrypted EBS volumes may expose sensitive data if the underlying hardware is compromised.",
            remediation_text="Create an encrypted snapshot and replace the unencrypted volume.",
            compliance="CIS-2.0 | 2.2.1 | NIST-800-53 | SC-28",
            categories="encryption | data-protection",
        ))

    # ═══════════════════════════════════════════════════════════
    # RDS checks
    # ═══════════════════════════════════════════════════════════

    # 19. RDS encryption — blank DESCRIPTION + RISK (community check style)
    rows.append(row(
        PROD,
        check_id="rds_instance_storage_encrypted",
        check_title="Ensure RDS instances are encrypted at rest",
        check_type="Software and Configuration Checks",
        status="FAIL",
        status_extended="RDS instance prod-postgres is not encrypted at rest.",
        muted="False",
        service_name="rds",
        severity="high",
        resource_type="AWSRDSDBInstance",
        resource_uid="arn:aws:rds:ap-southeast-2:111122223333:db:prod-postgres",
        resource_name="prod-postgres",
        region="ap-southeast-2",
        description=None,  # ← blank (Category 2)
        risk=None,         # ← blank (Category 2)
        remediation_text="Enable encryption at rest for the RDS instance by creating an encrypted read replica and promoting it.",
        compliance="CIS-2.0 | 2.3.1 | NIST-800-53 | SC-28",
        categories="encryption",
    ))

    # 20. RDS public access — row 1 (original)
    rows.append(row(
        PROD,
        check_id="rds_instance_no_public_access",
        check_title="Ensure RDS instances are not publicly accessible",
        check_type="Software and Configuration Checks",
        status="FAIL",
        status_extended="RDS instance prod-postgres is publicly accessible.",
        muted="False",
        service_name="rds",
        severity="critical",
        resource_type="AWSRDSDBInstance",
        resource_uid="arn:aws:rds:ap-southeast-2:111122223333:db:prod-postgres",
        resource_name="prod-postgres",
        region="ap-southeast-2",
        description="Checks whether RDS DB instances are publicly accessible.",
        risk="Publicly accessible RDS instances can be reached from the internet, increasing the attack surface.",
        remediation_text="Modify the RDS instance to disable public accessibility.",
        remediation_cli="aws rds modify-db-instance --db-instance-identifier prod-postgres --no-publicly-accessible",
        compliance="CIS-2.0 | 2.3.2",
        categories="internet-exposed | network-security",
    ))

    # 21. RDS public access — row 2 — EXACT DUPLICATE (dedup test)
    rows.append(row(
        PROD,
        check_id="rds_instance_no_public_access",
        check_title="Ensure RDS instances are not publicly accessible",
        check_type="Software and Configuration Checks",
        status="FAIL",
        status_extended="RDS instance prod-postgres is publicly accessible.",
        muted="False",
        service_name="rds",
        severity="critical",
        resource_type="AWSRDSDBInstance",
        resource_uid="arn:aws:rds:ap-southeast-2:111122223333:db:prod-postgres",  # ← same ARN
        resource_name="prod-postgres",
        region="ap-southeast-2",
        description="Checks whether RDS DB instances are publicly accessible.",
        risk="Publicly accessible RDS instances can be reached from the internet.",
        remediation_text="Modify the RDS instance to disable public accessibility.",
        compliance="CIS-2.0 | 2.3.2",
        categories="internet-exposed",
    ))

    # ═══════════════════════════════════════════════════════════
    # VPC / Networking
    # ═══════════════════════════════════════════════════════════

    # 22. VPC flow logs — medium + internet-exposed (likelihood override test)
    rows.append(row(
        PROD,
        check_id="vpc_flow_logs_enabled",
        check_title="Ensure VPC flow logging is enabled in all VPCs",
        check_type="Software and Configuration Checks",
        status="FAIL",
        status_extended="VPC vpc-0prod5678 does not have flow logging enabled.",
        muted="False",
        service_name="vpc",
        severity="medium",
        resource_type="AWSVPC",
        resource_uid="arn:aws:ec2:ap-southeast-2:111122223333:vpc/vpc-0prod5678",
        resource_name="vpc-0prod5678",
        region="ap-southeast-2",
        description="Checks whether VPC flow logs are enabled for each VPC.",
        risk="Without flow logs, network traffic is not captured and security incidents cannot be investigated.",
        remediation_text="Enable VPC flow logs for all VPCs and send logs to CloudWatch Logs or S3.",
        compliance="CIS-2.0 | 3.9 | NIST-800-53 | AU-12",
        categories="logging | internet-exposed",
    ))

    # ═══════════════════════════════════════════════════════════
    # CloudTrail / Logging
    # ═══════════════════════════════════════════════════════════

    # 23. CloudTrail multi-region — blank REMEDIATION_CODE_* (normal)
    rows.append(row(
        PROD,
        check_id="cloudtrail_multi_region_enabled",
        check_title="Ensure CloudTrail is enabled in all regions",
        check_type="Software and Configuration Checks",
        status="FAIL",
        status_extended="CloudTrail trail management-trail is not enabled in all regions.",
        muted="False",
        service_name="cloudtrail",
        severity="high",
        resource_type="AWSCloudTrailTrail",
        resource_uid="arn:aws:cloudtrail:ap-southeast-2:111122223333:trail/management-trail",
        resource_name="management-trail",
        region="ap-southeast-2",
        description="Checks whether CloudTrail is enabled in all AWS regions.",
        risk="Without multi-region CloudTrail, API calls in non-logged regions go undetected.",
        related_url="https://docs.aws.amazon.com/awscloudtrail/latest/userguide/receive-cloudtrail-log-files-from-multiple-regions.html",
        remediation_text="Update the CloudTrail trail to be multi-region.",
        remediation_url="https://docs.aws.amazon.com/awscloudtrail/latest/userguide/",
        remediation_nativeiac=None,       # ← blank (normal)
        remediation_terraform=None,       # ← blank (normal)
        remediation_cli="aws cloudtrail update-trail --name management-trail --is-multi-region-trail",
        remediation_other=None,           # ← blank (normal)
        compliance="CIS-2.0 | 3.1 | NIST-800-53 | AU-2",
        categories="logging | auditing",
    ))

    # 24. CloudWatch log group retention — informational severity
    rows.append(row(
        PROD,
        check_id="cloudwatch_log_group_retention_policy_specific_days_enabled",
        check_title="Ensure CloudWatch log groups have retention policies set",
        check_type="Software and Configuration Checks",
        status="FAIL",
        status_extended="Log group /aws/lambda/prod-api has no retention policy.",
        muted="False",
        service_name="cloudwatch",
        severity="informational",
        resource_type="AWSCloudWatchLogGroup",
        resource_uid="arn:aws:logs:ap-southeast-2:111122223333:log-group:/aws/lambda/prod-api",
        resource_name="/aws/lambda/prod-api",
        region="ap-southeast-2",
        description="Checks whether CloudWatch log groups have a retention period set.",
        risk="Without retention policies, log groups accumulate data indefinitely increasing storage costs.",
        remediation_text="Set a retention policy on the CloudWatch log group.",
        remediation_cli="aws logs put-retention-policy --log-group-name /aws/lambda/prod-api --retention-in-days 90",
        compliance=None,   # ← no compliance mapping for informational
        categories="logging | cost-optimisation",
    ))

    # ═══════════════════════════════════════════════════════════
    # Detection / Monitoring
    # ═══════════════════════════════════════════════════════════

    # 25. GuardDuty — MANUAL status
    rows.append(row(
        PROD,
        check_id="guardduty_is_enabled",
        check_title="Ensure GuardDuty is enabled",
        check_type="Software and Configuration Checks",
        status="MANUAL",
        status_extended="Manually verify GuardDuty is enabled and findings are being reviewed.",
        muted="False",
        service_name="guardduty",
        severity="high",
        resource_type="AWSGuardDutyDetector",
        resource_uid=None,
        resource_name="guardduty",
        region="ap-southeast-2",
        description=None,  # ← blank on MANUAL checks (common)
        risk="GuardDuty provides threat detection. Without it, malicious activity may go undetected.",
        remediation_text="Enable GuardDuty in all regions and configure findings notifications.",
        compliance="CIS-2.0 | 4.1 | NIST-800-53 | SI-3",
        categories="threat-detection",
    ))

    # 26. Config recorder — account-level, no region, no resource
    r = row(
        PROD,
        check_id="config_recorder_all_regions_enabled",
        check_title="Ensure AWS Config is enabled in all regions",
        check_type="Software and Configuration Checks",
        status="FAIL",
        status_extended="AWS Config recorder is not enabled in one or more regions.",
        muted="False",
        service_name="account",
        severity="medium",
        resource_type="AWSAccount",
        resource_uid=None,
        resource_name=None,
        region=None,
        description="Checks whether AWS Config is enabled and recording in all regions.",
        risk="Without AWS Config, resource configuration changes cannot be tracked or audited.",
        remediation_text="Enable AWS Config in all regions with an S3 delivery channel.",
        compliance="CIS-2.0 | 3.5 | NIST-800-53 | CM-8",
        categories="configuration | auditing",
    )
    r["RESOURCE_UID"] = None
    r["RESOURCE_NAME"] = None
    rows.append(r)

    # 27. SecurityHub — unknown future column
    r = row(
        PROD,
        check_id="securityhub_enabled",
        check_title="Ensure AWS Security Hub is enabled",
        check_type="Software and Configuration Checks",
        status="FAIL",
        status_extended="Security Hub is not enabled in ap-southeast-2.",
        muted="False",
        service_name="securityhub",
        severity="medium",
        resource_type="AWSSecurityHub",
        resource_uid=None,
        resource_name="security-hub",
        region="ap-southeast-2",
        description="Checks whether AWS Security Hub is enabled.",
        risk="Security Hub aggregates findings from multiple AWS services. Without it, security posture visibility is reduced.",
        remediation_text="Enable AWS Security Hub in all regions.",
        compliance="CIS-2.0 | 4.16",
        categories="configuration | monitoring",
    )
    r["FUTURE_PROWLER_COLUMN"] = "experimental-value-v5"  # ← unknown future column
    rows.append(r)

    # ═══════════════════════════════════════════════════════════
    # LEGACY account rows  (no Org UID/Name)
    # ═══════════════════════════════════════════════════════════

    # 28. LEGACY IAM MFA — global, no REGION, different account
    rows.append(row(
        LEGACY,
        check_id="iam_user_mfa_enabled_console_access",
        check_title="Ensure MFA is enabled for all IAM users with console access",
        check_type="Software and Configuration Checks",
        status="FAIL",
        status_extended="IAM user admin.legacy has console access but no MFA.",
        muted="False",
        service_name="iam",
        severity="critical",
        resource_type="AWSIAMUser",
        resource_uid="arn:aws:iam::777788889999:user/admin.legacy",
        resource_name="admin.legacy",
        region=None,
        description=None,  # ← blank (same community check pattern)
        risk=None,
        remediation_text="Enable MFA for the IAM user.",
        compliance="CIS-2.0 | 1.10",
        categories="identity-management | mfa",
    ))

    # 29. LEGACY EC2 SSH — same check_id as row 14, different account
    rows.append(row(
        LEGACY,
        check_id="ec2_securitygroup_allow_ingress_from_internet_to_tcp_port_22",
        check_title="Ensure no security groups allow ingress from 0.0.0.0/0 to port 22",
        check_type="Software and Configuration Checks",
        status="FAIL",
        status_extended="Security group sg-0legacy99 allows unrestricted SSH.",
        muted="False",
        service_name="ec2",
        severity="critical",
        resource_type="AWSEC2SecurityGroup",
        resource_uid="arn:aws:ec2:us-east-1:777788889999:security-group/sg-0legacy99",
        resource_name="sg-0legacy99",
        region="us-east-1",  # different region from PROD
        description="Checks whether security groups allow unrestricted inbound access on TCP port 22.",
        risk="Unrestricted SSH access allows brute-force login attempts.",
        remediation_text="Restrict SSH ingress to specific IP ranges.",
        compliance="CIS-2.0 | 5.2",
        categories="internet-exposed | network-security",
    ))

    # 30. LEGACY — completely blank optional fields stress test
    rows.append(row(
        LEGACY,
        check_id="s3_bucket_default_encryption",
        check_title="Ensure S3 bucket has default encryption enabled",
        check_type="Software and Configuration Checks",
        status="FAIL",
        status_extended=None,       # ← blank
        muted="False",
        service_name="s3",
        severity="medium",
        resource_type="AWSS3Bucket",
        resource_uid="arn:aws:s3:::legacy-archive",
        resource_name="legacy-archive",
        region="us-east-1",
        description=None,           # ← blank (Category 2)
        risk=None,                  # ← blank (Category 2)
        related_url=None,
        remediation_text=None,      # ← blank (Category 2)
        remediation_url=None,
        compliance=None,            # ← blank
        categories=None,            # ← blank
        resource_tags=None,
    ))

    return rows


def generate(output_path: str | Path) -> Path:
    """Write the messy synthetic dataset to an XLSX file."""
    output_path = Path(output_path)
    rows_data = build_rows()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "prowler-output"

    # Determine all columns (standard + any extras)
    all_cols = list(HEADERS)
    extra_cols = []
    for r in rows_data:
        for k in r:
            if k not in all_cols and k not in extra_cols:
                extra_cols.append(k)
    all_cols.extend(extra_cols)

    # Header row — bold
    for col_idx, col_name in enumerate(all_cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True)

    # Data rows
    for row_idx, row_dict in enumerate(rows_data, 2):
        for col_idx, col_name in enumerate(all_cols, 1):
            value = row_dict.get(col_name)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            # Force formula-like strings to text type
            if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
                cell.data_type = "s"

    wb.save(str(output_path))

    # ── Summary report ──
    status_counts: dict[str, int] = {}
    blank_desc = blank_risk = blank_remed = 0
    for r in rows_data:
        s = r.get("STATUS", "?")
        status_counts[s] = status_counts.get(s, 0) + 1
        if not r.get("DESCRIPTION"):
            blank_desc += 1
        if not r.get("RISK"):
            blank_risk += 1
        if not r.get("REMEDIATION_RECOMMENDATION_TEXT"):
            blank_remed += 1

    check_ids = sorted({r["CHECK_ID"] for r in rows_data})
    accounts = sorted({r["ACCOUNT_NAME"] for r in rows_data})

    print(f"\nMessy synthetic dataset written: {output_path}")
    print(f"  Total rows  : {len(rows_data)}")
    print(f"  Distinct checks: {len(check_ids)}")
    print(f"  Accounts    : {accounts}")
    print(f"  Status breakdown:")
    for s, c in sorted(status_counts.items()):
        print(f"    {s:20s}: {c}")
    print(f"  Blank DESCRIPTION      : {blank_desc}/{len(rows_data)}")
    print(f"  Blank RISK             : {blank_risk}/{len(rows_data)}")
    print(f"  Blank REMEDIATION_TEXT : {blank_remed}/{len(rows_data)}")
    print(f"  Extra columns          : {extra_cols}")

    return output_path


if __name__ == "__main__":
    out = Path(__file__).parent.parent / "data" / "synthetic" / "synthetic_prowler_messy.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    generate(out)
